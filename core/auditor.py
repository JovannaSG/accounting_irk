"""
Ядро автоаудита бухгалтерских баз 1С:Бухгалтерия (без 1С-интеграции).

Реализует 5 контрольных точек из ТЗ:
  4.1 Красное сальдо
  4.2 Развернутое сальдо
  4.3 Незакрытое сальдо на конец месяца
  4.4 Остатки на счете 000
  4.5 Незакрытые расчеты с контрагентами
"""

from __future__ import annotations

from collections.abc import Mapping
import io
from dataclasses import dataclass
from typing import Any, Sequence

import pandas as pd

from core import ml
from core import nlp
from core.formatting import fmt_date as _fmt_date
from core.formatting import fmt_num as _fmt_num
from core.formatting import fmt_rub as _fmt_rub
from core.formatting import period_sort_series

COLUMN_ALIASES: dict[str, list[str]] = {
    "Счет": ["Счет", "Счёт", "account", "schet"],
    "Субконто": ["Субконто", "sub_account", "Аналитика"],
    "Тип": ["Тип", "type"],
    "Период": ["Период", "period"],
    "Организация": ["Организация", "organization", "company"],
    "НачалоДебет": ["НачалоДебет", "begin_debit", "НД", "start_debit"],
    "НачалоКредит": ["НачалоКредит", "begin_credit", "НК", "start_credit"],
    "ОборотДебет": ["ОборотДебет", "debit_turnover", "ОД"],
    "ОборотКредит": ["ОборотКредит", "credit_turnover", "ОК"],
    "КонецДебет": ["КонецДебет", "end_debit", "СальдоКонецДебет", "КД", "Дебет"],
    "КонецКредит": ["КонецКредит", "end_credit", "СальдоКонецКредит", "КК", "Кредит"],
}

DOCUMENT_ALIASES: dict[str, list[str]] = {
    "Дата": ["Дата", "date"],
    "Документ": ["Документ", "document"],
    "Контрагент": ["Контрагент", "counterparty", "Контрагенты"],
    "Счет": ["Счет", "Счёт", "account"],
    "Вид": ["Вид", "type", "ВидОперации"],
    "Сумма": ["Сумма", "amount", "sum"],
    "Назначение": ["Назначение", "Назначение платежа", "НазначениеПлатежа", "purpose"],
}

REQUIRED_OSV: list[str] = ["Счет", "Тип", "КонецДебет", "КонецКредит"]
NUMERIC_OSV: list[str] = [
    "НачалоДебет", "НачалоКредит",
    "ОборотДебет", "ОборотКредит",
    "КонецДебет", "КонецКредит"
]
VALID_TYPES: set[str] = {"A", "P", "AP"}

EPS: float = 1e-6

OSV_COLUMNS: list[str] = [
    "Период", "Счет", "Субконто", "Тип", "Организация", "Договор",
    "НачалоДебет", "НачалоКредит", "ОборотДебет", "ОборотКредит",
    "КонецДебет", "КонецКредит",
]

DEFAULT_CLOSING_ACCOUNTS: list[str] = ["25", "26", "44", "90", "91", "99"]
SETTLEMENT_GROUPS: list[str] = ["60", "62", "76"]

CHECK_KEYS: dict[str, str] = {
    "red_balance": "4.1 Красное сальдо",
    "expanded_balance": "4.2 Развернутое сальдо",
    "unclosed_month_end": "4.3 Незакрытое сальдо на конец месяца",
    "account_000": "4.4 Счет 000",
    "settlements": "4.5 Незакрытые расчеты с контрагентами",
}

RECOMMENDATIONS: dict[str, str] = {
    "Красное сальдо: активный счет с кредитовым остатком": "Проверьте проводки по счету и корректность начальных остатков.",
    "Красное сальдо: пассивный счет с дебетовым остатком": "Проверьте проводки по счету. Дебетовый остаток на пассивном счете указывает на переплату или ошибку.",
    "Развернутое сальдо по аналитике": "У одного контрагента одновременно дебетовый и кредитовый остаток. Проверьте зачет аванса.",
    "Незакрытое сальдо на конец месяца (закрываемые счета)": "Выполните регламентную операцию «Закрытие месяца».",
    "Зависшее сальдо (не меняется между периодами)": "Проверьте, не «повисли» ли расчеты, которые должны были закрыться.",
    "Незакрытое сальдо на счете 000": "Служебный счет 000 должен быть закрыт. Проверьте корректность ввода остатков.",
    "Контрагенты: аванс и долг одновременно по разным счетам (ОСВ)": "У контрагента одновременно долг и аванс. Рекомендуется провести зачет аванса.",
    "Контрагенты: развернутое сальдо на счетах расчетов (без реестра документов)": "По контрагенту есть дебетовый и кредитовый остаток.",
    "Контрагенты: расчеты не закрыты документами": "По контрагенту остались незакрытые расчеты: непогашенный долг или переплата.",
    "Контрагенты: расхождение документов и остатков ОСВ": "Сумма документов не сходится с остатком ОСВ.",
    "ML: нетипичная сумма операции": "Сумма операции сильно отклоняется от обычных сумм по контрагенту.",
    "ML: резкий скачок оборотов между периодами": "Обороты по счету резко изменились.",
    "ML: возможные дубли контрагентов": "Обнаружены похожие названия контрагентов. Рекомендуется объединить дубли.",
    "NLP: подозрительные назначения платежей (115-ФЗ)": "Назначение платежа содержит формулировки с повышенным риском.",
    "Контроль групп счетов: незакрытые остатки": "Группа счетов не закрыта на конец периода.",
}

GROUP_PRESETS: dict[str, list[str]] = {
    "Авансы выданные/полученные": ["60.02", "62.02", "76.АВ", "76.ВА"],
    "Расходы будущих периодов": ["97"],
    "Товары": ["41", "43"],
    "Денежные средства": ["50", "51", "52", "55", "57", "58"],
    "Кредиты и займы": ["66", "67"],
}

DETAIL_COLUMNS: list[str] = [
    "Проверка", "Уровень", "Период", "Организация",
    "Счет", "Субконто", "Договор",
    "Дебет", "Кредит",
    "Сумма", "Комментарий"
]

_LEVEL_ORDER: dict[str, int] = {"warning": 1, "error": 2}
_LEVEL_RU: dict[str, str] = {"error": "Ошибка", "warning": "Предупреждение"}

SECTION_SPECS: list[tuple[str, tuple[str, ...]]] = [
    ("Закрытие месяца", (
        "Незакрытое сальдо на конец месяца (закрываемые счета)",
        "Зависшее сальдо (не меняется между периодами)",
        "Незакрытое сальдо на счете 000",
        "Контроль групп счетов: незакрытые остатки",
    )),
    ("Красное и развернутое сальдо", (
        "Красное сальдо: активный счет с кредитовым остатком",
        "Красное сальдо: пассивный счет с дебетовым остатком",
        "Развернутое сальдо по аналитике",
    )),
    ("Расчеты с контрагентами", (
        "Контрагенты: аванс и долг одновременно по разным счетам (ОСВ)",
        "Контрагенты: развернутое сальдо на счетах расчетов (без реестра документов)",
        "Контрагенты: расчеты не закрыты документами",
        "Контрагенты: расхождение документов и остатков ОСВ",
    )),
    ("ML: нетипичные суммы операций", ("ML: нетипичная сумма операции",)),
    ("ML: скачки оборотов между периодами", ("ML: резкий скачок оборотов между периодами",)),
    ("ML: дубли контрагентов", ("ML: возможные дубли контрагентов",)),
    ("NLP: подозрительные назначения платежей (115-ФЗ)", ("NLP: подозрительные назначения платежей (115-ФЗ)",)),
]

_RUB_PDF_COLUMNS: set[str] = {
    "Дебет", "Кредит", "ОборотДебет", "ОборотКредит",
    "Сумма", "Медиана", "НачалоДебет", "НачалоКредит",
}
_NUM_PDF_COLUMNS: set[str] = {"Отклонение", "Отношение", "Сходство"}
_LEVEL_PDF_COLOR: dict[str, tuple[int, int, int]] = {
    "error": (156, 0, 6),
    "warning": (156, 101, 0),
}

_PDF_COL_STYLE: dict[str, tuple[float, str]] = {
    "Проверка": (32, "LEFT"),
    "Уровень": (14, "LEFT"),
    "Период": (16, "LEFT"),
    "Организация": (20, "LEFT"),
    "Счет": (14, "LEFT"),
    "Субконто": (26, "LEFT"),
    "Дата": (16, "LEFT"),
    "Документ": (20, "LEFT"),
    "Контрагент": (24, "LEFT"),
    "Вид": (13, "LEFT"),
    "Дебет": (16, "RIGHT"),
    "Кредит": (16, "RIGHT"),
    "ОборотДебет": (17, "RIGHT"),
    "ОборотКредит": (17, "RIGHT"),
    "Сумма": (18, "RIGHT"),
    "Медиана": (15, "RIGHT"),
    "Отклонение": (15, "RIGHT"),
    "Отношение": (15, "RIGHT"),
    "Сходство": (15, "RIGHT"),
    "Название А": (24, "LEFT"),
    "Название Б": (24, "LEFT"),
    "Комментарий": (46, "LEFT"),
}

_MONEY_COLUMNS: set[str] = {
    "Дебет", "Кредит", "Сумма", "Медиана",
    "НачалоДебет", "НачалоКредит",
    "ОборотДебет", "ОборотКредит",
}
_EXCEL_MONEY_FORMAT = "#,##0.00"

TOP_FINDINGS_LIMIT = 10
AUDIT_LOGIC_VERSION = "1.2"

_SHORT_PDF_LABELS: dict[str, str] = {
    "Незакрытое сальдо на конец месяца (закрываемые счета)": "Сальдо не закрыто",
    "Зависшее сальдо (не меняется между периодами)": "Зависшее сальдо",
    "Незакрытое сальдо на счете 000": "Счет 000",
    "Контроль групп счетов: незакрытые остатки": "Группа не закрыта",
    "Красное сальдо: активный счет с кредитовым остатком": "Красное сальдо (А)",
    "Красное сальдо: пассивный счет с дебетовым остатком": "Красное сальдо (П)",
    "Развернутое сальдо по аналитике": "Развернутое сальдо",
    "Контрагенты: аванс и долг одновременно по разным счетам (ОСВ)": "Аванс и долг",
    "Контрагенты: развернутое сальдо на счетах расчетов (без реестра документов)": "Аванс и долг (без реестра)",
    "Контрагенты: расчеты не закрыты документами": "Расчеты не закрыты",
    "Контрагенты: расхождение документов и остатков ОСВ": "Расхождение с ОСВ",
}


@dataclass
class Finding(Mapping):
    level: str
    title: str
    data: pd.DataFrame
    amount: float

    def __getitem__(self, key: str) -> Any:
        return getattr(self, key)

    def __iter__(self):
        return iter(["level", "title", "data", "amount"])

    def __len__(self) -> int:
        return 4


def account_group(code: object) -> str:
    return str(code).strip().split(".")[0]


def _group_account_string(accounts_str: object) -> str:
    """
    Переводит строку '60.01, 62.02' в '60, 62' (родительские счета без дублей)
    """

    if pd.isna(accounts_str) or not str(accounts_str).strip():
        return ""

    codes: list = []
    raw_codes = str(accounts_str).replace(";", ",").split(",")
    for c in raw_codes:
        stripped_c = c.strip()
        # Только не пустые строки добавляем
        if stripped_c:
            codes.append(stripped_c)

    unique_groups = set()
    for c in codes:
        parent_account = account_group(c)
        unique_groups.add(parent_account)

    grouped = sorted(unique_groups)
    return ", ".join(grouped)


def _rename_by_aliases(df: pd.DataFrame, aliases: dict) -> pd.DataFrame:
    rename: dict = {}
    for canonical, names in aliases.items():
        for name in names:
            if name in df.columns:
                rename[name] = canonical
                break
    return df.rename(columns=rename)


def normalize_balances(df: pd.DataFrame) -> pd.DataFrame:
    df = _rename_by_aliases(df, COLUMN_ALIASES).copy()

    missing = [c for c in REQUIRED_OSV if c not in df.columns]
    if missing:
        raise ValueError(f"Отсутствуют обязательные колонки: {', '.join(missing)}")

    df["Счет"] = df["Счет"].astype(str).str.strip()
    df["Тип"] = df["Тип"].astype(str).str.strip().str.upper()
    df["Субконто"] = (
        df["Субконто"].fillna("-")
        if "Субконто" in df.columns
        else pd.Series("-", index=df.index)
    ).astype(str)
    df["Договор"] = (
        df["Договор"].fillna('-')
        if "Договор" in df.columns
        else pd.Series('-', index=df.index)
    ).astype(str)
    df["Период"] = (
        df["Период"].fillna("")
        if "Период" in df.columns
        else pd.Series("", index=df.index)
    ).astype(str).str.strip()
    df["Организация"] = (
        df["Организация"].fillna('-')
        if "Организация" in df.columns
        else pd.Series('-', index=df.index)
    ).astype(str)

    bad_types = set(df["Тип"]) - VALID_TYPES
    if bad_types:
        raise ValueError(f"Недопустимые значения Тип: {sorted(bad_types)}")

    for col in NUMERIC_OSV:
        if col in df.columns:
            clean_series = df[col].astype(str).str.strip().replace(["-", ""], "0")
            coerced = pd.to_numeric(clean_series, errors="coerce")
            invalid = clean_series.notna() & (clean_series != "nan") & coerced.isna()
            if invalid.any():
                bad = clean_series.loc[invalid].dropna().unique()[:5]
                raise ValueError(f"Колонка '{col}' содержит нечисловые значения: {list(bad)}")
            df[col] = coerced.fillna(0.0)
        else:
            df[col] = 0.0

    if df.empty:
        raise ValueError("Файл ОСВ пуст")

    return df[[c for c in OSV_COLUMNS if c in df.columns]]


def normalize_documents(df: pd.DataFrame) -> pd.DataFrame:
    df = _rename_by_aliases(df, DOCUMENT_ALIASES).copy()

    missing: list[str] = [
        c
        for c in ["Дата", "Контрагент", "Вид", "Сумма"]
        if c not in df.columns
    ]
    if missing:
        raise ValueError(
            f"В файле документов отсутствуют колонки: {', '.join(missing)}"
        )

    df["Контрагент"] = df["Контрагент"].astype(str).str.strip()
    df["Дата"] = pd.to_datetime(df["Дата"], errors="coerce")
    df["Вид"] = df["Вид"].astype(str).str.strip().str.lower()
    df["Сумма"] = pd.to_numeric(df["Сумма"], errors="coerce")

    if "Назначение" in df.columns:
        df["Назначение"] = df["Назначение"].fillna("").astype(str).str.strip()

    if df["Сумма"].isna().any():
        raise ValueError("Колонка 'Сумма' содержит нечисловые значения")

    kind: dict[str, str] = {
        "отгрузка": "отгрузка", "реализация": "отгрузка",
        "продажа": "отгрузка", "оплата": "оплата",
        "платеж": "оплата", "платёж": "оплата",
        "аванс": "аванс", "предоплата": "аванс",
    }
    df["ВидНорм"] = df["Вид"].map(kind)
    df.loc[df["ВидНорм"].isna(), "ВидНорм"] = df.loc[df["ВидНорм"].isna(), "Вид"]

    if df.empty:
        raise ValueError("Файл документов пуст")

    return df


class AutoAuditor1C:
    def __init__(
        self,
        balances_df: pd.DataFrame,
        documents_df: pd.DataFrame | None = None,
        closing_accounts: list | None = None,
        checks: set[str] | None = None,
        meta: dict | None = None,
        balance_group_checks: bool = False,
        stuck_balance_checks: bool = False,
        ml_enabled: bool = False,
        nlp_enabled: bool = True,
        ml_amount_anomalies: bool = True,
        ml_turnover_jumps: bool = True,
        ml_duplicates: bool = True,
        anomaly_k: float = ml.DEFAULT_K,
        anomaly_min_abs: float = ml.DEFAULT_MIN_ABS,
        anomaly_min_ops: int = ml.DEFAULT_MIN_OPS,
        jump_ratio: float = ml.DEFAULT_JUMP_RATIO,
        jump_min_abs: float = ml.DEFAULT_JUMP_MIN_ABS,
        dup_threshold: int = ml.DEFAULT_SIM_THRESHOLD,
    ) -> None:
        self.balances = normalize_balances(balances_df)
        if documents_df is not None:
            self.documents = normalize_documents(documents_df)
        else:
            self.documents = None
        self.closing_accounts: set[str] = {
            str(a).split(".")[0]
            for a in (closing_accounts or DEFAULT_CLOSING_ACCOUNTS)
        }
        self.ml_enabled = ml_enabled
        self.ml_amount_anomalies = ml_amount_anomalies
        self.ml_turnover_jumps = ml_turnover_jumps
        self.ml_duplicates = ml_duplicates
        self.nlp_enabled = nlp_enabled
        self.anomaly_k = anomaly_k
        self.anomaly_min_abs = anomaly_min_abs
        self.anomaly_min_ops = anomaly_min_ops
        self.jump_ratio = jump_ratio
        self.jump_min_abs = jump_min_abs
        self.dup_threshold = dup_threshold
        self.checks: set[str] | None = checks
        if checks is not None:
            unknown = set(checks) - set(CHECK_KEYS)
            if unknown:
                raise ValueError(f"Неизвестные ключи: {', '.join(sorted(unknown))}")
        self.balance_group_checks = balance_group_checks
        self.stuck_balance_checks = stuck_balance_checks
        self.meta: dict = meta or {}
        self.errors: list[Finding] = []

    def _check_enabled(self, key: str) -> bool:
        if not CHECK_KEYS.get(key):
            raise ValueError(f"Неизвестный ключ: {key}")
        return self.checks is None or key in self.checks

    def _add(self, level: str, title: str, data: pd.DataFrame) -> None:
        """
        Перехватчик: считает модули ошибок,
        переводит субсчета в счета и суммирует без сальдирования
        """

        if data.empty:
            return

        data = data.copy()

        # Сначала считаем сумму ошибки для каждой отдельной строки (до группировки)
        # Это гарантирует, что ошибки не "съедят" друг друга
        # (60.01 Д 45к + 60.02 К 30к = 75к, а не 15к)
        if "Сумма" not in data.columns:
            if {"КонецДебет", "КонецКредит"} <= set(data.columns):
                if "Развернутое" in title:
                    data["Сумма"] = data["КонецДебет"] + data["КонецКредит"]
                else:
                    data["Сумма"] = (data["КонецДебет"] - data["КонецКредит"]).abs()
            else:
                data["Сумма"] = 0.0

        # Убираем прошлые месяцы и переводим субсчета в родительские (60.01 -> 60)
        # Только для OSV/балансовых проверок. ML/NLP-проверки в этот набор не входят:
        # их находки — отдельные события, субсчета не схлопываем
        balance_checks: set = {
            "Красное сальдо: активный счет с кредитовым остатком",
            "Красное сальдо: пассивный счет с дебетовым остатком",
            "Развернутое сальдо по аналитике",
            "Незакрытое сальдо на конец месяца (закрываемые счета)",
            "Зависшее сальдо (не меняется между периодами)",
            "Незакрытое сальдо на счете 000",
            "Контроль групп счетов: незакрытые остатки",
            "Контрагенты: аванс и долг одновременно по разным счетам (ОСВ)",
            "Контрагенты: развернутое сальдо на счетах расчетов (без реестра документов)",
            "Контрагенты: расчеты не закрыты документами",
            "Контрагенты: расхождение документов и остатков ОСВ",
        }

        if title in balance_checks:
            # Убираем прошлые месяцы ПО ИСХОДНЫМ субсчетам (до схлопывания)
            # Сортируем по хронологии
            # (Период может быть в разных форматах: 31.01.2026 / 2026-01-31 / Январь),
            # иначе лексикографическая сортировка ломает границу годов
            if "Период" in data.columns:
                data["_pkey"] = period_sort_series(data["Период"])
                data = data.sort_values("_pkey", kind="stable", na_position="last")
                dedup_cols = [
                    c
                    for c in ["Организация", "Счет", "Субконто", "Договор"]
                    if c in data.columns
                ]
                data = data.drop_duplicates(subset=dedup_cols, keep="last")
                data = data.drop(columns="_pkey")

            # Добавляем в комментарий ТОЧНЫЙ субсчет (58.03), чтобы после схлопывания
            # в счёт (58) бухгалтер сразу видел, где именно спрятана ошибка
            if "Счет" in data.columns and "Комментарий" in data.columns:
                is_subaccount = data["Счет"].astype(str).str.contains(r"\.")
                if is_subaccount.any():
                    for idx in data[is_subaccount].index:
                        sub_acc = str(data.loc[idx, "Счет"]).strip()
                        comment = str(data.loc[idx, "Комментарий"])
                        # Защита от дублирования: субсчет уже упомянут в тексте
                        # (напр. «60.01 Д ...») — не вставляем повторно
                        if (
                            f"на субсчете {sub_acc}" not in comment
                            and f"{sub_acc} " not in comment
                        ):
                            data.loc[idx, "Комментарий"] = f"{comment} (на субсчете {sub_acc})"

            # Заменяем субсчета на родительские (60.01 -> 60)
            if "Счет" in data.columns:
                data["Счет"] = data["Счет"].apply(_group_account_string)

            # Группируем по родительскому счету и складываем модули сумм (Сумма)
            group_cols = [
                c
                for c in ["Период", "Организация", "Счет", "Субконто", "Договор"]
                if c in data.columns
            ]

            agg_funcs: dict = {}
            for col in data.columns:
                if col not in group_cols:
                    if col in {"КонецДебет", "КонецКредит", "Сумма"}:
                        agg_funcs[col] = "sum"
                    elif col == "Комментарий":
                        # Разные субсчета одного родителя могут иметь разные тексты ошибок —
                        # объединяем их через " | " (порядок сохраняем, дубли убираем)
                        agg_funcs[col] = lambda x: " | ".join(
                            dict.fromkeys(str(v) for v in x if pd.notna(v))
                        )
                    else:
                        agg_funcs[col] = "first"

            data = data.groupby(group_cols, dropna=False, as_index=False).agg(agg_funcs)

        if data.empty:
            return

        # Итоговая сумма для карточки ошибки (уже корректно сложенные модули)
        if "Сумма" in data.columns:
            amount = float(data["Сумма"].fillna(0.0).sum())
        else:
            amount = 0.0

        self.errors.append(Finding(
            level=level, title=title,
            data=data, amount=amount
        ))

    @staticmethod
    def _first_occurrence(
        b: pd.DataFrame,
        keys: list[str],
        flag: pd.Series
    ) -> dict[tuple, str]:
        b2 = b.loc[flag, keys + ["Период"]].copy()
        b2["_sort"] = pd.to_datetime(
            b2["Период"],
            errors="coerce",
            format="mixed"
        )
        b2 = b2.sort_values(keys + ["_sort"])
        first = b2.groupby(keys, as_index=False)["Период"].first()
        return {
            tuple(row[k] for k in keys): row["Период"]
            for _, row in first.iterrows()
            if row["Период"]
        }

    def _annotate_since(
        self,
        b: pd.DataFrame,
        flag: pd.Series,
        base_comment: str,
        since_text: str
    ) -> pd.DataFrame:
        sub = b[flag].copy()
        if sub.empty:
            return sub

        since = self._first_occurrence(
            b,
            ["Организация", "Счет", "Субконто"],
            flag
        )
        keys = pd.Series(
            list(zip(sub["Организация"], sub["Счет"], sub["Субконто"])),
            index=sub.index
        )
        periods = keys.map(since)

        sub["Комментарий"] = base_comment
        mask_has_period = periods.notna()
        if mask_has_period.any():
            sub.loc[mask_has_period, "Комментарий"] = (
                base_comment + "; "
                + since_text + ' '
                + periods[mask_has_period].map(_fmt_date)
            )
        return sub

    def check_red_balance(self) -> None:
        b = self.balances
        if b.empty:
            return

        net = b["КонецДебет"] - b["КонецКредит"]

        # Активные счета с кредитовым сальдо
        active = self._annotate_since(
            b, (b["Тип"] == "A") & (net < -EPS),
            "Активный счет имеет кредитовое (отрицательное) сальдо",
            "отрицательное сальдо с"
        )
        if not active.empty:
            self._add(
                "error",
                "Красное сальдо: активный счет с кредитовым остатком",
                active
            )

        # Пассивные счета с дебетовым сальдо
        passive = self._annotate_since(
            b, (b["Тип"] == "P") & (net > EPS),
            "Пассивный счет имеет дебетовое (отрицательное) сальдо",
            "отрицательное сальдо с"
        )
        if not passive.empty:
            self._add(
                "error",
                "Красное сальдо: пассивный счет с дебетовым остатком",
                passive
            )

    def check_expanded_balance(self) -> None:
        b = self.balances
        both = b[
            (b["Субконто"] != "-")
            & (b["КонецДебет"] > 0)
            & (b["КонецКредит"] > 0)
        ].copy()
        both["Комментарий"] = "По контрагенту/аналитике одновременно \
            дебетовое и кредитовое сальдо"
        self._add("warning", "Развернутое сальдо по аналитике", both)

    def check_unclosed_month_end(self) -> None:
        b = self.balances

        # Отбираем только строки с закрываемыми счетами
        closing_mask = b["Счет"].map(account_group).isin(self.closing_accounts)
        closing_df = b[closing_mask].copy()

        if not closing_df.empty:
            # Для закрываемых счетов важен итог по РОДИТЕЛЬСКОМУ счету: субсчета
            # (90.01 Выручка, 90.09 Прибыль/убыток) копят остаток весь год, а к нулю
            # должны сходиться именно родители 90/91. Сворачиваем субсчета в родителя
            closing_df["Счет"] = closing_df["Счет"].map(account_group)
            closing_df["Субконто"] = "-"
            closing_df["Договор"] = "-"

            grp_cols: list = ["Период", "Организация", "Счет", "Субконто", "Договор"]
            parent_b = closing_df.groupby(
                grp_cols,
                dropna=False,
                as_index=False
            )[["КонецДебет", "КонецКредит"]].sum()

            # Сальдо родительского счета = |Дебет - Кредит| (по всем субсчетам разом)
            parent_b["Сумма"] = (parent_b["КонецДебет"] - parent_b["КонецКредит"]).abs()

            # Аннотируем датой первого появления (по истории периодов родителя).
            unclosed_parents = self._annotate_since(
                parent_b, parent_b["Сумма"] > EPS,
                "Остаток по закрываемому счету в целом",
                "остаток с"
            )

            if not unclosed_parents.empty:
                # Корректируем Д/К: показываем весь остаток на той стороне, где он есть.
                unclosed_parents["КонецДебет"] = unclosed_parents.apply(
                    lambda r:
                        r["Сумма"]
                        if r["КонецДебет"] > r["КонецКредит"]
                        else 0.0, axis=1
                )
                unclosed_parents["КонецКредит"] = unclosed_parents.apply(
                    lambda r: r["Сумма"]
                        if r["КонецКредит"] > r["КонецДебет"]
                        else 0.0, axis=1
                )

                self._add(
                    "error",
                    "Незакрытое сальдо на конец месяца (закрываемые счета)",
                    unclosed_parents
                )

        # Зависшее сальдо (stuck_balance_checks) остаётся без изменений: тут важно
        # отслеживать движение каждого отдельного субсчета
        if not self.stuck_balance_checks:
            return
        b_nonzero = b[(b["КонецДебет"] > 0) | (b["КонецКредит"] > 0)]
        if b_nonzero.empty:
            return

        b2 = b_nonzero.copy()
        b2["_pkey"] = period_sort_series(b2["Период"])

        b2 = b2.sort_values(
            ["Организация", "Счет", "Субконто", "_pkey"],
            kind="stable",
            na_position="last"
        )
        grp = b2.groupby(["Организация", "Счет", "Субконто"], sort=False)

        prev_d = grp["КонецДебет"].shift(1)
        prev_k = grp["КонецКредит"].shift(1)

        stuck_mask = (
            ((b2["КонецДебет"] - prev_d).abs() < EPS)
            & ((b2["КонецКредит"] - prev_k).abs() < EPS)
            & prev_d.notna()
        )
        if not stuck_mask.any():
            return

        b2["_stuck"] = stuck_mask.to_numpy()
        b2["_prev_period"] = b2["Период"].shift(1)
        prev_stuck = b2["_stuck"].shift(1, fill_value=False)

        same_group = (
            (b2["Организация"] == b2["Организация"].shift(1))
            & (b2["Счет"] == b2["Счет"].shift(1))
            & (b2["Субконто"] == b2["Субконто"].shift(1))
        )
        streak_start = b2["_stuck"] & ~(prev_stuck & same_group)
        b2["_streak"] = streak_start.cumsum()

        rows: list[dict[str, Any]] = []
        value_cols = [c for c in b2.columns if not c.startswith("_")]
        for _, streak in b2[b2["_stuck"]].groupby("_streak", sort=False):
            last = streak.iloc[-1]
            n_periods = len(streak) + 1
            row = {c: last[c] for c in value_cols}
            row["Комментарий"] = f"Сальдо не меняется с {_fmt_date(streak.iloc[0]['_prev_period'])} ({n_periods} пер.)"
            rows.append(row)

        self._add(
            "warning",
            "Зависшее сальдо (не меняется между периодами)",
            pd.DataFrame(rows)
        )

    @staticmethod
    def _matches_group_preset(code: Any, preset: list[str]) -> bool:
        code = str(code)
        for p in preset:
            if "." in p:
                if code == p:
                    return True
            elif account_group(code) == p:
                return True
        return False

    def check_group_balances(self) -> None:
        if not self.balance_group_checks:
            return

        b = self.balances
        periods = sorted(set(b["Период"]))
        if not periods:
            return

        last_period = periods[-1]
        rows: list = []
        for group_name, preset in GROUP_PRESETS.items():
            matched = b[
                b["Счет"].map(
                    lambda code: self._matches_group_preset(code, preset)
                )
            ]
            if matched.empty:
                continue

            parents = matched[matched["Субконто"] == "-"]
            use = parents if not parents.empty else matched
            g = use[use["Период"] == last_period]
            if g.empty:
                continue

            g_unique = g.drop_duplicates(
                subset=["Организация", "Счет", "Субконто"]
            )
            d = float(g_unique["КонецДебет"].sum())
            k = float(g_unique["КонецКредит"].sum())

            net = d - k
            if abs(net) <= EPS:
                continue

            if "Организация" in g_unique.columns:
                org = g_unique["Организация"].iloc[0]
            else:
                org = "-"

            if net > 0:
                indicator = "Д"
            else:
                indicator = "К"
            formatted_rub = _fmt_rub(abs(net))

            rows.append({
                "Период": last_period,
                "Организация": org,
                "Счет": ", ".join(preset),
                "Субконто": group_name,
                "КонецДебет": d,
                "КонецКредит": k,
                "Сумма": abs(net),
                "Комментарий": f"Группа «{group_name}» не закрыта: остаток {indicator} {formatted_rub}",
            })
        if rows:
            self._add(
                "warning",
                "Контроль групп счетов: незакрытые остатки",
                pd.DataFrame(rows)
            )

    def check_account_000(self) -> None:
        acc = self.balances[
            (self.balances["Счет"].map(account_group) == "000")
            & ((self.balances["КонецДебет"] > 0) | (self.balances["КонецКредит"] > 0))
        ].copy()
        acc["Комментарий"] = "Незакрытый остаток на служебном счете 000"
        self._add("error", "Незакрытое сальдо на счете 000", acc)

    def _osv_settlement_balance(self) -> pd.Series:
        b = self.balances
        sett = b[b["Счет"].map(account_group).isin(SETTLEMENT_GROUPS)]
        return sett.groupby("Субконто")["КонецДебет"].sum() - \
            sett.groupby("Субконто")["КонецКредит"].sum()

    def _osv_settlement_breakdown(self, subconto: Any) -> dict[str, float]:
        b = self.balances
        sett = b[
            (b["Счет"].map(account_group).isin(SETTLEMENT_GROUPS))
            & (b["Субконто"] == subconto)
        ]
        out: dict[str, float] = {}
        for account, g in sett.groupby("Счет"):
            net = float(g["КонецДебет"].sum() - g["КонецКредит"].sum())
            if abs(net) > EPS:
                out[str(account)] = net
        return out

    def _reference_date(self) -> pd.Timestamp:
        dates = pd.to_datetime(self.balances["Период"], errors="coerce").dropna()
        if not dates.empty:
            return dates.max()
        if self.documents is not None:
            dates = pd.to_datetime(self.documents["Дата"], errors="coerce").dropna()
            if not dates.empty:
                return dates.max()
        return pd.Timestamp.today()

    @staticmethod
    def _oldest_unconsumed(
        rows: list[tuple],
        consume: float
    ) -> pd.Timestamp | None:
        rows = [r for r in rows if pd.notna(r[0]) and r[1] > 0]
        rows.sort(key=lambda r: r[0])
        remaining = consume
        for d, amount in rows:
            if remaining < amount - EPS:
                return pd.Timestamp(d)
            remaining -= amount
        return None

    @staticmethod
    def _settlement_components(
        shipped: float,
        paid: float,
        advances: float
    ) -> tuple[float, float, float, float]:
        open_amount = shipped - paid - advances
        debt = max(open_amount, 0.0)
        credit_side = max(-open_amount, 0.0)
        advance_left = min(advances, credit_side)
        overpaid = credit_side - advance_left
        consumed = advances - advance_left
        return debt, advance_left, overpaid, consumed

    def _check_settlement_advance_vs_debt(self) -> None:
        b = self.balances
        sett = b[
            (b["Счет"].map(account_group).isin(SETTLEMENT_GROUPS))
            & (b["Субконто"] != "-")
        ]
        rows: list = []

        # Группируем не только по Субконто (контрагенту), но и по ДОГОВОРУ:
        # аванс по одному договору и долг по другому у одного контрагента — это норма
        grp_cols: list[str] = ["Период", "Организация", "Субконто", "Договор"]

        for keys, g in sett.groupby(grp_cols):
            period, org, subconto, contract = keys

            for group in ("60", "62"):
                gs = g[g["Счет"].map(account_group) == group]
                if gs.empty:
                    continue

                nets: dict = {}
                for account, gg in gs.groupby("Счет"):
                    account_str = str(account)
                    net_value = float(gg["КонецДебет"].sum() - gg["КонецКредит"].sum())
                    nets[account_str] = net_value

                debit_sum = 0.0
                credit_sum = 0.0
                for n in nets.values():
                    if n > EPS:
                        debit_sum += n
                    elif n < -EPS:
                        credit_sum += -n

                # Если на одном и том же договоре есть и Дт,
                # и Кт — ошибка (не зачтен аванс)
                if debit_sum <= EPS or credit_sum <= EPS:
                    continue

                parts: list = []
                for a, n in sorted(nets.items()):
                    indicator = "Д" if n > 0 else "К"
                    formatted_rub = _fmt_rub(abs(n))
                    parts.append(f"{a} {indicator} {formatted_rub}")

                rows.append({
                    "Период": period,
                    "Организация": org,
                    "Счет": ", ".join(sorted(nets)),
                    "Субконто": subconto,
                    "Договор": contract,
                    "КонецДебет": debit_sum,
                    "КонецКредит": credit_sum,
                    "Сумма": debit_sum + credit_sum,
                    "Комментарий": f"Долг и аванс на разных счетах по одному договору: {'; '.join(parts)}",
                })

        if rows:
            self._add(
                "warning",
                "Контрагенты: аванс и долг одновременно по разным счетам (ОСВ)",
                pd.DataFrame(rows)
            )

    def check_unclosed_settlements(self) -> None:
        self._check_settlement_advance_vs_debt()

        if self.documents is None:
            sett = self.balances[self.balances["Счет"].map(account_group).isin(SETTLEMENT_GROUPS)]
            dup = sett[(sett["Субконто"] != "-") & (sett["КонецДебет"] > 0) & (sett["КонецКредит"] > 0)].copy()
            dup["Комментарий"] = "Возможные незакрытые расчеты: аванс и долг по одному контрагенту"
            self._add("warning", "Контрагенты: развернутое сальдо на счетах расчетов (без реестра документов)", dup)
            return

        rows: list = []
        for name, g in self.documents.groupby("Контрагент"):
            g = g.sort_values("Дата")
            shipped = float(g.loc[g["ВидНорм"] == "отгрузка", "Сумма"].sum())
            paid = float(g.loc[g["ВидНорм"] == "оплата", "Сумма"].sum())
            advances = float(g.loc[g["ВидНорм"] == "аванс", "Сумма"].sum())

            debt, advance_left, overpaid, consumed = self._settlement_components(shipped, paid, advances)
            open_amount = debt - advance_left - overpaid

            issues: list[str] = []
            if debt > EPS:
                issues.append(f"остаток долга {_fmt_rub(debt)}")
            if advance_left > EPS:
                issues.append(f"незачтенный аванс {_fmt_rub(advance_left)}")
            if overpaid > EPS:
                issues.append(f"переплата {_fmt_rub(overpaid)}")

            if debt > EPS:
                oldest = self._oldest_unconsumed(
                    list(g.loc[
                        g["ВидНорм"] == "отгрузка",
                        ["Дата", "Сумма"]
                    ].itertuples(index=False, name=None)),
                    paid + consumed,
                )
                if oldest is not None:
                    issues.append(f"старейший долг от {_fmt_date(oldest)}")
            if advance_left > EPS:
                oldest = self._oldest_unconsumed(
                    list(g.loc[
                        g["ВидНорм"] == "аванс",
                        ["Дата", "Сумма"]
                    ].itertuples(index=False, name=None)),
                    consumed,
                )
                if oldest is not None:
                    issues.append(f"старейший незачтенный аванс от {_fmt_date(oldest)}")

            osv_break = self._osv_settlement_breakdown(name)
            if osv_break:
                accounts = ", ".join(sorted(osv_break))
            else:
                accounts = ", ".join(sorted(set(g["Счет"].astype(str))))

            comment = "; ".join(issues) if issues else "Расчеты закрыты"
            osv_net = float(self._osv_settlement_balance().get(name, 0.0))
            if abs(osv_net - open_amount) > 1e-6:
                comment += f"; расхождение с ОСВ ({_fmt_rub(osv_net)})"

            rows.append({
                "Период": "",
                "Организация": "-",
                "Счет": accounts,
                "Субконто": name,
                "КонецДебет": max(open_amount, 0.0),
                "КонецКредит": max(-open_amount, 0.0),
                "Сумма": abs(open_amount),
                "Комментарий": comment,
            })

        res = pd.DataFrame(rows)
        if not res.empty:
            problems = res[res["Сумма"] > EPS]
            self._add(
                "error",
                "Контрагенты: расчеты не закрыты документами",
                problems.copy()
            )
            mismatches = res[res["Комментарий"].str.contains("расхождение", na=False)]
            self._add(
                "warning",
                "Контрагенты: расхождение документов и остатков ОСВ",
                mismatches.copy()
            )

    def check_amount_anomalies(self) -> None:
        if self.documents is None:
            return
        found = ml.detect_amount_anomalies(
            self.documents, k=self.anomaly_k,
            min_abs=self.anomaly_min_abs, min_ops=self.anomaly_min_ops
        )
        self._add("warning", "ML: нетипичная сумма операции", found)

    def check_turnover_jumps(self) -> None:
        found = ml.detect_turnover_jumps(
            self.balances, ratio=self.jump_ratio,
            min_abs=self.jump_min_abs
        )
        self._add(
            "warning",
            "ML: резкий скачок оборотов между периодами",
            found
        )

    def check_duplicate_counterparties(self) -> None:
        names: list[object] = []
        if "Субконто" in self.balances.columns:
            names += list(self.balances["Субконто"].dropna())
        if (
            self.documents is not None
            and "Контрагент" in self.documents.columns
        ):
            names += list(self.documents["Контрагент"].dropna())
        found = ml.find_duplicate_counterparties(
            names,
            threshold=self.dup_threshold
        )
        self._add("warning", "ML: возможные дубли контрагентов", found)

    def run_ml_checks(self) -> None:
        if not self.ml_enabled:
            return
        if self.ml_amount_anomalies:
            self.check_amount_anomalies()
        if self.ml_turnover_jumps:
            self.check_turnover_jumps()
        if self.ml_duplicates:
            self.check_duplicate_counterparties()

    def check_payment_purpose_risks(self) -> None:
        if self.documents is None:
            return
        found = nlp.detect_payment_risks(self.documents)
        self._add(
            "warning",
            "NLP: подозрительные назначения платежей (115-ФЗ)",
            found
        )

    def run_audit(self) -> list[Finding]:
        self.errors = []
        if self._check_enabled("red_balance"):
            self.check_red_balance()
        if self._check_enabled("expanded_balance"):
            self.check_expanded_balance()
        if self._check_enabled("unclosed_month_end"):
            self.check_unclosed_month_end()
            self.check_group_balances()
        if self._check_enabled("account_000"):
            self.check_account_000()
        if self._check_enabled("settlements"):
            self.check_unclosed_settlements()
        self.run_ml_checks()
        if self.nlp_enabled:
            self.check_payment_purpose_risks()
        return self.errors

    @classmethod
    def from_findings(
        cls,
        errors: list[dict],
        meta: dict | None = None
    ) -> "AutoAuditor1C":
        instance = cls.__new__(cls)
        instance.balances = pd.DataFrame(columns=OSV_COLUMNS)
        instance.documents = None
        instance.meta = dict(meta or {})
        findings: list[Finding] = []
        for e in errors or []:
            data = e.get("data")
            if not isinstance(data, pd.DataFrame):
                data = pd.DataFrame(data or [])
            findings.append(Finding(
                level=str(e.get("level", "warning")),
                title=str(e.get("title", "")),
                data=data,
                amount=float(e.get("amount") or 0.0),
            ))
        instance.errors = findings
        return instance

    def summary_df(self) -> pd.DataFrame:
        columns: list[str] = [
            "Проверка", "Уровень",
            "Период", "Строк",
            "Сумма", "Рекомендации"
        ]
        details = self.details_df()
        if details.empty:
            return pd.DataFrame(columns=columns)
        grouped = (
            details
            .groupby(["Проверка", "Уровень", "Период"], sort=False)
            .agg(Строк=("Сумма", "count"), Сумма=("Сумма", "sum"))
            .reset_index()
        )
        grouped["Рекомендации"] = grouped["Проверка"].map(RECOMMENDATIONS).fillna("")
        return grouped[columns]

    def details_df(self) -> pd.DataFrame:
        rows: list = []
        for e in self.errors:
            for _, r in e["data"].iterrows():
                rows.append({
                    "Проверка": e["title"],
                    "Уровень": e["level"],
                    "Период": r.get("Период", ""),
                    "Организация": str(r.get("Организация", "")),
                    "Счет": str(r.get("Счет", "")),  # Счета уже сгруппированы в _add
                    "Субконто": r.get("Субконто", ""),
                    "Договор": r.get("Договор", ""),
                    "Дебет": r.get("КонецДебет", 0.0),
                    "Кредит": r.get("КонецКредит", 0.0),
                    "Сумма": (
                        float(r["Сумма"])
                        if "Сумма" in r.index
                        else abs(
                            float(r.get("КонецДебет", 0.0) or 0.0)
                            - float(r.get("КонецКредит", 0.0) or 0.0)
                        )
                    ),
                    "Комментарий": r.get("Комментарий", ""),
                })
        return pd.DataFrame(rows, columns=DETAIL_COLUMNS)

    def top_findings_df(self, limit: int = 10) -> pd.DataFrame:
        columns: list[str] = [
            "Проверка", "Уровень",
            "Период", "Организация",
            "Счет", "Субконто",
            "Сумма"
        ]
        details = self.details_df()
        if details.empty or limit <= 0:
            return pd.DataFrame(columns=columns)
        ranked = details[columns].copy()
        ranked["_abs"] = ranked["Сумма"].abs()
        ranked = ranked.sort_values("_abs", ascending=False).head(limit)
        return ranked.drop(columns="_abs").reset_index(drop=True)

    def _sections(self) -> list[dict[str, Any]]:
        used: set[str] = set()
        sections: list[dict[str, Any]] = []

        specs = list(SECTION_SPECS)

        details = self.details_df()
        if details.empty:
            return []

        for sec_title, titles in specs:
            mask = details["Проверка"].isin(titles)
            if not mask.any():
                continue

            sec_df = details[mask].copy()
            used.update(titles)

            if len(titles) > 1:
                sec_df["Проверка"] = sec_df["Проверка"].map(
                    lambda x: _SHORT_PDF_LABELS.get(str(x), str(x))
                )
            else:
                sec_df = sec_df.drop(columns=["Проверка"], errors="ignore")

            recs = " ".join(
                dict.fromkeys(
                    RECOMMENDATIONS.get(t, "")
                    for t in titles
                    if RECOMMENDATIONS.get(t, "")
                )
            )
            if "error" in sec_df["Уровень"].unique():
                level = "error"
            else:
                level = "warning"
            sec_df = self._drop_empty_columns(sec_df)

            sections.append({
                "title": sec_title,
                "level": level,
                "recommendation": recs,
                "df": sec_df,
            })

        remaining_mask = ~details["Проверка"].isin(used)
        if remaining_mask.any():
            rem_df = details[remaining_mask]
            for title, group in rem_df.groupby("Проверка"):
                grp = group.drop(columns=["Проверка"], errors="ignore")
                if "error" in grp["Уровень"].unique():
                    level = "error"
                else:
                    level = "warning"
                sections.append({
                    "title": str(title),
                    "level": level,
                    "recommendation": RECOMMENDATIONS.get(str(title), ""),
                    "df": self._drop_empty_columns(grp),
                })
        return sections

    @staticmethod
    def _drop_empty_columns(df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return df
        keep: list[str] = []
        for col in df.columns:
            series = df[col]
            non_empty = series.notna() & (
                series.astype(str).str.strip().ne("")
                & series.astype(str).ne("nan")
            )
            if bool(non_empty.any()):
                keep.append(col)
        return df[keep]

    @staticmethod
    def _account_codes_in(cell: object) -> list[str]:
        return [
            c.strip()
            for c in str(cell).replace(";", ",").split(",")
            if c.strip()
        ]

    def accounts_with_errors(self) -> list[str]:
        codes: set[str] = set()
        details = self.details_df()
        if not details.empty and "Счет" in details.columns:
            for cell in details["Счет"].dropna():
                codes.update(self._account_codes_in(cell))
        return sorted(codes)

    def account_subaccounts(self, account_code: str) -> list[str]:
        code = str(account_code).strip()
        if "." in code:
            return [code]
        # родительский счет сам по себе (после схлопывания 60.01 -> 60)
        sub: set[str] = {code}
        if self.balances is not None and "Счет" in self.balances.columns:
            values = self.balances["Счет"].dropna()
            idx = 0
            while idx < len(values):
                item = str(values.iloc[idx]).strip()
                if account_group(item) == code:
                    sub.add(item)
                idx += 1
        return sorted(sub)

    def account_report_df(self, account_code: str) -> pd.DataFrame:
        details = self.details_df()
        if details.empty:
            return details
        subaccounts = set(self.account_subaccounts(account_code))
        return details[
            details["Счет"].map(
                lambda cell: bool(
                    set(self._account_codes_in(cell))
                    & subaccounts
                )
            )
        ]

    def account_subconto(self, account_code: str) -> list[str]:
        names: set[str] = set()
        subaccounts = self.account_subaccounts(account_code)
        if (
            "Счет" in self.balances.columns
            and "Субконто" in self.balances.columns
        ):
            rows = self.balances[self.balances["Счет"].isin(subaccounts)]["Субконто"]
            names.update(
                str(n).strip()
                for n in rows.dropna()
                if str(n).strip() != "-"
            )
        if (
            self.documents is not None
            and "Счет" in self.documents.columns
            and "Контрагент" in self.documents.columns
        ):
            rows = self.documents[self.documents["Счет"].isin(subaccounts)]["Контрагент"]
            names.update(
                str(n).strip()
                for n in rows.dropna()
                if str(n).strip()
            )
        return sorted(names)

    def account_subconto_duplicates(
        self,
        account_code: str,
        threshold: int | None = None
    ) -> pd.DataFrame:
        names = self.account_subconto(account_code)
        if not names:
            return pd.DataFrame(columns=[
                "Субконто", "Название А",
                "Название Б", "Сходство",
                "Комментарий"
            ])
        return ml.find_duplicate_counterparties(
            names,
            threshold=threshold if threshold is not None else self.dup_threshold
        )

    def accounts_summary_df(self) -> pd.DataFrame:
        columns: list[str] = [
            "Счет", "Кол-во нарушений",
            "Проверки", "Периоды",
            "Сумма", "Дубли контрагентов"
        ]

        details = self.details_df()
        if details.empty:
            return pd.DataFrame(columns=columns)

        rows: list = []
        for account in self.accounts_with_errors():
            g = self.account_report_df(account)
            checks = sorted(set(g["Проверка"].astype(str)))
            periods = sorted(p for p in set(g["Период"].astype(str)) if p)

            # Мы не суммируем ошибку 12 раз! Берем только актуальную сумму
            max_per_issue = g.groupby(
                ["Проверка", "Организация", "Субконто"],
                dropna=False
            )["Сумма"].max()
            total_sum = round(float(max_per_issue.sum()), 2)

            rows.append({
                "Счет": account,
                "Кол-во нарушений": len(g),
                "Проверки": "; ".join(checks),
                "Периоды": ", ".join(periods),
                "Сумма": total_sum,
                "Дубли контрагентов": len(self.account_subconto_duplicates(account)),
            })
        return pd.DataFrame(rows, columns=columns)

    def _meta_payload(self) -> dict:
        payload: dict = {}
        for key in ("title", "organization", "period"):
            value = self.meta.get(key)
            if value:
                payload[key] = value
        return payload

    def report(self) -> dict:
        details = self.details_df()
        if not details.empty:
            total_flags = len(details)
            total_amount = float(
                details.groupby(
                    ["Проверка", "Организация", "Счет", "Субконто"],
                    dropna=False
                )["Сумма"].max().sum()
            )
        else:
            total_flags = 0
            total_amount = 0.0

        result = {
            "total_flags": total_flags,
            "total_amount": total_amount,
            "summary": self.summary_df(),
            "details": details
        }
        result.update(self._meta_payload())
        if not self.errors:
            result.update(status="ok", status_label="Успешно")
        else:
            result.update(status="warning", status_label="Есть ошибки")
        return result

    def to_excel(self, account_pass: dict | None = None) -> bytes:
        from openpyxl.styles import Alignment, Font, PatternFill

        summary = self.summary_df()
        details = self.details_df()
        details = details.drop(
            columns=["Субконто", "Договор"],
            errors="ignore"
        )
        by_account = self.accounts_summary_df()
        top_findings = self.top_findings_df(TOP_FINDINGS_LIMIT)
        top_findings = top_findings.drop(
            columns=["Субконто"],
            errors="ignore"
        )

        pass_details = pd.DataFrame()
        if account_pass is not None:
            pd_ = account_pass.get("details_df")
            if pd_ is not None and not getattr(pd_, "empty", True):
                pass_details = pd_

        n_err = sum(len(e["data"]) for e in self.errors if e["level"] == "error")
        if n_err:
            status_label = "Есть ошибки"
        else:
            if self.errors:
                status_label = "Есть предупреждения"
            else:
                status_label = "Успешно"

        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            meta_payload = self._meta_payload()
            meta_rows = [
                {"Параметр": "Заголовок", "Значение": meta_payload.get("title", "")},
                {"Параметр": "Организация", "Значение": meta_payload.get("organization", "")},
                {"Параметр": "Период", "Значение": meta_payload.get("period", "")},
                {"Параметр": "Сформировано", "Значение": pd.Timestamp.now().strftime("%d.%m.%Y %H:%M")},
                {"Параметр": "Статус", "Значение": status_label},
            ]
            pd.DataFrame(meta_rows).to_excel(writer, sheet_name="Обзор", index=False)

            summary.to_excel(writer, sheet_name="Сводный отчет", index=False)
            details.to_excel(writer, sheet_name="Детальный отчет", index=False)
            by_account.to_excel(writer, sheet_name="По счетам", index=False)
            if not pass_details.empty:
                pass_details.to_excel(
                    writer,
                    sheet_name="Проход по счетам",
                    index=False
                )

            wb = writer.book
            header_fill = PatternFill("solid", fgColor="4472C4")
            header_font = Font(color="FFFFFF", bold=True)
            red_fill = PatternFill("solid", fgColor="FFC7CE")
            yellow_fill = PatternFill("solid", fgColor="FFEB9C")
            title_font = Font(bold=True, size=12)
            header_align = Alignment(
                horizontal="center",
                vertical="top",
                wrap_text=True
            )

            def _finish_table_sheet(ws) -> None:
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_align
                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions
                self._excel_fit_widths(ws)
                self._excel_money_formats(ws)

            overview = writer.sheets["Обзор"]
            blocks: list[tuple[int, pd.DataFrame]] = []

            def _add_block(
                cursor: int,
                title: str,
                frame: pd.DataFrame
            ) -> int:
                overview.cell(row=cursor + 1, column=1, value=title).font = title_font
                header_row = cursor + 2
                frame.to_excel(
                    writer,
                    sheet_name="Обзор",
                    index=False,
                    startrow=cursor + 1
                )
                blocks.append((header_row, frame))
                return cursor + 2 + len(frame) + 1

            cursor = len(meta_rows) + 2
            if not summary.empty:
                cursor = _add_block(cursor, "Сводка по проверкам", summary)
            if not top_findings.empty:
                cursor = _add_block(
                    cursor,
                    "Крупнейшие нарушения",
                    top_findings
                )

            overview["A1"].fill = header_fill
            overview["B1"].fill = header_fill
            overview["A1"].font = header_font
            overview["B1"].font = header_font
            for i in range(2, len(meta_rows) + 1):
                overview.cell(row=i, column=1).font = Font(bold=True)
            for header_row, frame in blocks:
                for j in range(1, len(frame.columns) + 1):
                    cell = overview.cell(row=header_row, column=j)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_align
                if "Уровень" in frame.columns:
                    lvl_j = list(frame.columns).index("Уровень") + 1
                    for i in range(len(frame)):
                        level = overview.cell(row=header_row + 1 + i, column=lvl_j).value
                        if level == "error":
                            fill = red_fill
                        else:
                            if level == "warning":
                                fill = yellow_fill
                            else:
                                fill = None
                        if fill:
                            for j in range(1, len(frame.columns) + 1):
                                overview.cell(row=header_row + 1 + i, column=j).fill = fill
                for j, h in enumerate(frame.columns, start=1):
                    if str(h) in _MONEY_COLUMNS:
                        for row_cells in overview.iter_rows(
                            min_row=header_row + 1,
                            max_row=header_row + len(frame),
                            min_col=j,
                            max_col=j
                        ):
                            for cell in row_cells:
                                cell.number_format = _EXCEL_MONEY_FORMAT

            sheets_to_color = [("Сводный отчет", 1), ("Детальный отчет", 1)]
            if not pass_details.empty:
                sheets_to_color.append(("Проход по счетам", 2))

            for sheet, level_col in sheets_to_color:
                ws = wb[sheet]
                _finish_table_sheet(ws)
                for row in ws.iter_rows(min_row=2):
                    level = row[level_col].value
                    if level == "error":
                        fill = red_fill
                    else:
                        if level == "warning":
                            fill = yellow_fill
                        else:
                            fill = None
                    if fill:
                        for cell in row:
                            cell.fill = fill

            _finish_table_sheet(wb["По счетам"])

        return buf.getvalue()

    @staticmethod
    def _excel_fit_widths(ws, cap: int = 55) -> None:
        from openpyxl.utils import get_column_letter
        widths: dict[int, int] = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                length = len(str(cell.value))
                if length > widths.get(cell.column, 0):
                    widths[cell.column] = length
        for idx, width in widths.items():
            ws.column_dimensions[get_column_letter(idx)].width = min(
                cap,
                max(9, width + 2)
            )

    @staticmethod
    def _excel_money_formats(ws) -> None:
        from openpyxl.utils import get_column_letter
        for cell in ws[1]:
            if str(cell.value) in _MONEY_COLUMNS:
                letter = get_column_letter(cell.column)
                for row_cells in ws[f"{letter}2:{letter}{ws.max_row}"]:
                    for c in row_cells:
                        c.number_format = _EXCEL_MONEY_FORMAT

    @staticmethod
    def _find_pdf_font() -> str | None:
        import os
        bundled = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "fonts", "DejaVuSans.ttf"
        )
        return bundled if os.path.exists(bundled) else None

    @staticmethod
    def _register_pdf_fonts(pdf, font_path: str) -> bool:
        import os as _os
        pdf.add_font("DejaVu", "", font_path)
        bold = _os.path.join(_os.path.dirname(font_path), "DejaVuSans-Bold.ttf")
        if _os.path.exists(bold):
            pdf.add_font("DejaVu", "B", bold)
            return True
        return False

    @staticmethod
    def _pdf_table(
        pdf,
        headers: Sequence[str],
        rows: Sequence[Sequence[Any]],
        widths: Sequence[float],
        has_bold: bool = False,
    ) -> None:
        from fpdf.fonts import FontFace
        usable = pdf.w - pdf.l_margin - pdf.r_margin
        total = float(sum(widths)) or 1.0
        col_widths = [w / total * usable for w in widths]

        aligns: list[str] = []
        fmts: list = []
        for h in headers:
            if h in _RUB_PDF_COLUMNS:
                aligns.append("RIGHT")
                fmts.append(_fmt_rub)
            elif h in _NUM_PDF_COLUMNS:
                aligns.append("RIGHT")
                fmts.append(_fmt_num)
            elif h in ("Дата", "Период"):
                aligns.append("LEFT")
                fmts.append(_fmt_date)
            else:
                aligns.append("LEFT")
                fmts.append(lambda v: "" if v is None else str(v))

        headings_style = FontFace(
            fill_color=(235, 238, 242),
            emphasis="BOLD" if has_bold else None
        )
        pdf.set_font("DejaVu", size=8)
        with pdf.table(
            col_widths=tuple(col_widths),
            text_align=tuple(aligns),
            line_height=4.4,
            padding=1.2,
            headings_style=headings_style,
            cell_fill_color=(249, 250, 251),
            cell_fill_mode="ROWS",
        ) as table:
            hr = table.row()
            for h in headers:
                hr.cell(h)
            for row in rows:
                tr = table.row()
                for i, v in enumerate(row):
                    tr.cell(fmts[i](v))
        pdf.ln(2)

    def to_pdf(self, account_pass: dict | None = None) -> bytes:
        try:
            from fpdf import FPDF
            from fpdf.enums import XPos, YPos
        except ImportError as exc:
            raise ImportError("PDF-экспорт требует библиотеку fpdf2.") from exc

        font_path = self._find_pdf_font()
        if font_path is None:
            raise RuntimeError("Не найден шрифт с кириллицей для PDF.")

        class _AuditPdf(FPDF):
            def footer(self) -> None:
                self.set_y(-12)
                self.set_font("DejaVu", size=8)
                self.set_text_color(130, 130, 130)
                self.cell(0, 6, f"Стр. {self.page_no()}", align="C")

        pdf = _AuditPdf(format="A4")
        pdf.set_margins(10, 10, 10)
        has_bold = self._register_pdf_fonts(pdf, font_path)
        pdf.set_auto_page_break(auto=True, margin=16)
        pdf.add_page()

        meta = self._meta_payload()
        pdf.set_font("DejaVu", size=14, style="B" if has_bold else "")
        pdf.cell(
            0, 8,
            meta.get("title", "Отчет аудита бухгалтерских баз 1С"),
            new_x=XPos.LMARGIN, new_y=YPos.NEXT
        )
        pdf.ln(2)
        pdf.set_font("DejaVu", size=10)
        pdf.cell(
            0, 6,
            f"Организация: {meta.get('organization', '—')}",
            new_x=XPos.LMARGIN, new_y=YPos.NEXT
        )
        pdf.cell(
            0, 6,
            f"Период: {meta.get('period', '—')}",
            new_x=XPos.LMARGIN, new_y=YPos.NEXT
        )
        pdf.cell(
            0, 6,
            f"Дата отчета: {pd.Timestamp.now().strftime('%d.%m.%Y %H:%M')}",
            new_x=XPos.LMARGIN, new_y=YPos.NEXT
        )

        sections = self._sections()
        n_err_rows = int(sum(
                len(e["data"])
                for e in self.errors
                if e["level"] == "error"
        ))
        n_warn_rows = int(sum(
            len(e["data"])
            for e in self.errors
            if e["level"] == "warning"
        ))
        counts = []
        if n_err_rows:
            counts.append(f"Ошибок: {n_err_rows}")
        if n_warn_rows:
            counts.append(f"Предупреждений: {n_warn_rows}")
        pdf.cell(
            0, 6,
            " · ".join(counts) if counts else "Нарушений не выявлено",
            new_x=XPos.LMARGIN, new_y=YPos.NEXT
        )

        if n_err_rows:
            status = "Есть ошибки"
            status_rgb = (156, 0, 6)
        else:
            if n_warn_rows:
                status = "Есть предупреждения"
                status_rgb = (156, 101, 0)
            else:
                status = "Успешно"
                status_rgb = (0, 128, 0)
        pdf.set_text_color(*status_rgb)
        pdf.set_font("DejaVu", size=10, style="B" if has_bold else "")
        pdf.cell(
            0, 6,
            f"Статус: {status}",
            new_x=XPos.LMARGIN, new_y=YPos.NEXT
        )
        pdf.set_text_color(0, 0, 0)

        if not sections:
            pdf.ln(4)
            pdf.set_text_color(0, 128, 0)
            pdf.cell(
                0, 7,
                "Нарушений не выявлено.",
                new_x=XPos.LMARGIN, new_y=YPos.NEXT
            )
            pdf.set_text_color(0, 0, 0)
            return bytes(pdf.output())

        pdf.ln(4)
        summary = self.summary_df()
        if not summary.empty:
            pdf.set_text_color(68, 114, 196)
            pdf.set_font("DejaVu", size=11, style="B" if has_bold else "")
            pdf.cell(
                0, 7,
                "Сводка по проверкам",
                new_x=XPos.LMARGIN, new_y=YPos.NEXT
            )
            pdf.set_text_color(0, 0, 0)
            sum_rows = [
                [
                    _SHORT_PDF_LABELS.get(str(r["Проверка"]), str(r["Проверка"])),
                    _LEVEL_RU.get(str(r["Уровень"]), str(r["Уровень"])),
                    int(r["Строк"]), float(r["Сумма"]),
                ]
                for _, r in summary.iterrows()
            ]
            self._pdf_table(
                pdf,
                ["Проверка", "Уровень", "Строк", "Сумма"],
                sum_rows,
                [64, 26, 14, 24],
                has_bold=has_bold
            )

            top = self.top_findings_df(TOP_FINDINGS_LIMIT)
            if not top.empty:
                pdf.ln(2)
                pdf.set_text_color(68, 114, 196)
                pdf.set_font("DejaVu", size=11, style="B" if has_bold else "")
                pdf.cell(
                    0, 7,
                    "Крупнейшие нарушения",
                    new_x=XPos.LMARGIN, new_y=YPos.NEXT
                )
                pdf.set_text_color(0, 0, 0)
                top_rows = [
                    [
                        _SHORT_PDF_LABELS.get(
                            str(r["Проверка"]), str(r["Проверка"])),
                            r["Период"], r["Организация"],
                            r["Счет"], r["Субконто"],
                            float(r["Сумма"]
                        ),
                    ]
                    for _, r in top.iterrows()
                ]
                self._pdf_table(
                    pdf,
                    ["Проверка", "Период", "Орг.", "Счет", "Субконто", "Сумма"],
                    top_rows,
                    [30, 16, 18, 12, 34, 20],
                    has_bold=has_bold,
                )

        for s in sections:
            df = s["df"]
            if pdf.get_y() > 230:
                pdf.add_page()
            pdf.ln(3)
            pdf.set_text_color(*_LEVEL_PDF_COLOR.get(s["level"], (0, 0, 0)))
            pdf.set_font("DejaVu", size=11, style="B" if has_bold else "")
            pdf.cell(0, 7, s["title"], new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            pdf.set_text_color(95, 95, 95)
            pdf.set_font("DejaVu", size=8.5)
            if s["recommendation"]:
                pdf.multi_cell(0, 4.2, s["recommendation"])
            pdf.set_text_color(0, 0, 0)
            pdf.ln(1)

            headers = list(df.columns)
            widths = [
                _PDF_COL_STYLE.get(h, (20, "LEFT"))[0]
                for h in headers
            ]
            rows = [[v for v in rec] for rec in df.to_numpy()]
            self._pdf_table(pdf, headers, rows, widths, has_bold=has_bold)

        pass_details = pd.DataFrame()
        pass_dups = pd.DataFrame()
        if account_pass is not None:
            details = account_pass.get("details_df")
            dups = account_pass.get("duplicates_df")
            if details is not None:
                pass_details = details
            if dups is not None:
                pass_dups = dups

        if not pass_details.empty:
            if pdf.get_y() > 200:
                pdf.add_page()
            else:
                pdf.ln(4)
            pdf.set_text_color(68, 114, 196)
            pdf.set_font("DejaVu", size=11, style="B" if has_bold else "")
            pdf.cell(
                0, 7,
                "Автопроход по счетам",
                new_x=XPos.LMARGIN, new_y=YPos.NEXT
            )
            pdf.set_text_color(0, 0, 0)

            pass_headers: list[str] = [
                "Проверка", "Период",
                "Субконто", "Дебет",
                "Кредит", "Сумма"
            ]
            pass_widths = [_PDF_COL_STYLE[h][0] for h in pass_headers]
            accounts = sorted(set(
                str(c)
                for c in pass_details["Счет"].dropna()
            ))
            for account in accounts:
                acc_rows = pass_details[pass_details["Счет"].astype(str) == account]
                dups_n = 0
                if not pass_dups.empty and "Счет" in pass_dups.columns:
                    dups_n = int((pass_dups["Счет"].astype(str) == account).sum())
                note = f"Счет {account} — нарушений: {len(acc_rows)}"
                if dups_n:
                    note += f", дублей контрагентов: {dups_n}"
                pdf.set_font("DejaVu", size=10, style="B" if has_bold else "")
                pdf.cell(0, 6, note, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                table_rows = [
                    [
                        r.get("Проверка", ""), r.get("Период", ""),
                        r.get("Субконто", ""), r.get("Дебет", 0.0),
                        r.get("Кредит", 0.0), r.get("Сумма", 0.0),
                    ]
                    for _, r in acc_rows.iterrows()
                ]
                self._pdf_table(
                    pdf,
                    pass_headers,
                    table_rows,
                    pass_widths,
                    has_bold=has_bold
                )

        return bytes(pdf.output())
