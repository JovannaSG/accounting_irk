"""
Ядро автоаудита бухгалтерских баз 1С:Бухгалтерия (без 1С-интеграции).

Реализует 5 контрольных точек из ТЗ:
  4.1 Красное сальдо
  4.2 Развернутое сальдо
  4.3 Незакрытое сальдо на конец месяца
  4.4 Остатки на счете 000
  4.5 Незакрытые расчеты с контрагентами

Входные данные (ОСВ):  Период, Счет, Субконто, Тип, НачалоДебет, НачалоКредит,
                        ОборотДебет, ОборотКредит, КонецДебет, КонецКредит
Опционально (документы): Дата, Документ, Контрагент, Счет, Вид, Сумма
"""

from __future__ import annotations

import io
from typing import Any, Optional

import pandas as pd

COLUMN_ALIASES: dict[str, list[str]] = {
    "Счет": ["Счет", "Счёт", "account", "schet"],
    "Субконто": ["Субконто", "sub_account", "Аналитика"],
    "Тип": ["Тип", "type"],
    "Период": ["Период", "period"],
    "НачалоДебет": ["НачалоДебет", "begin_debit", "НД", "start_debit"],
    "НачалоКредит": ["НачалоКредит", "begin_credit", "НК", "start_credit"],
    "ОборотДебет": ["ОборотДебет", "debit_turnover", "ОД"],
    "ОборотКредит": ["ОборотКредит", "credit_turnover", "ОК"],
    "КонецДебет": ["КонецДебет", "end_debit", "СальдоКонецДебет", "КД", "Дебет"],
    "КонецКредит": ["КонецКредит", "end_credit", "СальдоКонецКредит", "КК", "Кредит"],
}

DOCUMENT_ALIASES: dict[str, list[str]] = {
    "Дата": ["Дата", "date"],
    "Документ": ["Документ", "document"],
    "Контрагент": ["Контрагент", "counterparty", "Контрагенты"],
    "Счет": ["Счет", "Счёт", "account"],
    "Вид": ["Вид", "type", "ВидОперации"],
    "Сумма": ["Сумма", "amount", "sum"],
}

REQUIRED_OSV: list[str] = ["Счет", "Тип", "КонецДебет", "КонецКредит"]
NUMERIC_OSV: list[str] = [
    "НачалоДебет", "НачалоКредит",
    "ОборотДебет", "ОборотКредит",
    "КонецДебет", "КонецКредит"
]
VALID_TYPES: set[str] = {"A", "P", "AP"}

DEFAULT_CLOSING_ACCOUNTS: list[str] = ["25", "26", "44", "90", "91", "99"]
SETTLEMENT_GROUPS: list[str] = ["60", "62", "76"]

DETAIL_COLUMNS: list[str] = [
    "Проверка", "Уровень", "Период", 
    "Счет", "Субконто",
    "Дебет", "Кредит",
    "Сумма", "Комментарий"
]


def account_group(code: object) -> str:
    """
    Группа счета: '60.01' -> '60'; '000' -> '000'
    """
    return str(code).strip().split(".")[0]


def _rename_by_aliases(df: pd.DataFrame, aliases: dict) -> pd.DataFrame:
    rename: dict = {}
    for canonical, names in aliases.items():
        for name in names:
            if name in df.columns:
                rename[name] = canonical
                break
    return df.rename(columns=rename)


def normalize_balances(df: pd.DataFrame) -> pd.DataFrame:
    """
    Приводит ОСВ к канонической схеме и типам
    """

    df = _rename_by_aliases(df, COLUMN_ALIASES).copy()

    missing = [c for c in REQUIRED_OSV if c not in df.columns]
    if missing:
        raise ValueError(f"Отсутствуют обязательные колонки: {', '.join(missing)}")

    df["Счет"] = df["Счет"].astype(str).str.strip()
    df["Тип"] = df["Тип"].astype(str).str.strip().str.upper()
    df["Субконто"] = (
        df["Субконто"].fillna("-")
        if "Субконто" in df.columns
        else pd.Series("-", index=df.index)
    ).astype(str)
    df["Период"] = (
        df["Период"].fillna("")
        if "Период" in df.columns
        else pd.Series("", index=df.index)
    ).astype(str).str.strip()

    bad_types = set(df["Тип"]) - VALID_TYPES
    if bad_types:
        raise ValueError(f"Недопустимые значения Тип: {sorted(bad_types)}. Ожидаются: A, P, AP")

    for col in NUMERIC_OSV:
        if col in df.columns:
            coerced = pd.to_numeric(df[col], errors="coerce")
            invalid = df[col].notna() & coerced.isna()
            if invalid.any():
                bad = df.loc[invalid, col].dropna().unique()[:5]
                raise ValueError(f"Колонка '{col}' содержит нечисловые значения: {list(bad)}")
            df[col] = coerced.fillna(0.0)
        else:
            df[col] = 0.0

    if df.empty:
        raise ValueError("Файл ОСВ пуст")

    order: list[str] = [
        "Период", "Счет", "Субконто",
        "Тип", "НачалоДебет", "НачалоКредит",
        "ОборотДебет", "ОборотКредит", "КонецДебет",
        "КонецКредит"
    ]
    return df[[c for c in order if c in df.columns]]


def normalize_documents(df: pd.DataFrame) -> pd.DataFrame:
    """
    Приводит реестр документов к канонической схеме
    """

    df = _rename_by_aliases(df, DOCUMENT_ALIASES).copy()

    missing = [c for c in ["Дата", "Контрагент", "Вид", "Сумма"] if c not in df.columns]
    if missing:
        raise ValueError(f"В файле документов отсутствуют колонки: {', '.join(missing)}")

    df["Контрагент"] = df["Контрагент"].astype(str).str.strip()
    df["Дата"] = pd.to_datetime(df["Дата"], errors="coerce")
    df["Вид"] = df["Вид"].astype(str).str.strip().str.lower()
    df["Сумма"] = pd.to_numeric(df["Сумма"], errors="coerce")

    if df["Сумма"].isna().any():
        raise ValueError("Колонка 'Сумма' в документах содержит нечисловые значения")
    df["Сумма"] = df["Сумма"].fillna(0.0)

    kind: dict[str, str] = {
        "отгрузка": "отгрузка",
        "реализация": "отгрузка",
        "продажа": "отгрузка",
        "оплата": "оплата",
        "платеж": "оплата",
        "платёж": "оплата",
        "аванс": "аванс",
        "предоплата": "аванс",
    }
    df["ВидНорм"] = df["Вид"].map(kind)
    df.loc[df["ВидНорм"].isna(), "ВидНорм"] = df.loc[df["ВидНорм"].isna(), "Вид"]

    if df.empty:
        raise ValueError("Файл документов пуст")

    return df


class AutoAuditor1C:
    """
    Выполняет контрольные проверки ОСВ по правилам из ТЗ
    """

    def __init__(
        self,
        balances_df: pd.DataFrame,
        documents_df: Optional[pd.DataFrame] = None,
        closing_accounts: Optional[list] = None
    ) -> None:
        self.balances = normalize_balances(balances_df)
        self.documents = normalize_documents(documents_df) if documents_df is not None else None
        self.closing_accounts: set[str] = {
            str(a).split(".")[0]
            for a in (closing_accounts or DEFAULT_CLOSING_ACCOUNTS)
        }
        self.errors: list[dict] = []

    # ============ Вспомогательные методы ============
    def _add(self, level: str, title: str, data: pd.DataFrame) -> None:
        if data.empty:
            return
        amount = float((data["КонецДебет"] - data["КонецКредит"]).abs().sum())
        self.errors.append({
            "level": level,
            "title": title,
            "data": data,
            "amount": amount
        })

    # ============ 4.1 Красное сальдо ============
    def check_red_balance(self) -> None:
        b = self.balances
        net = b["КонецДебет"] - b["КонецКредит"]
        active = b[(b["Тип"] == "A") & (net < 0)]
        passive = b[(b["Тип"] == "P") & (net > 0)]
        if not active.empty:
            active["Комментарий"] = "Активный счет имеет кредитовое (отрицательное) сальдо"
            self._add("error", "Красное сальдо: активный счет с кредитовым остатком", active)
        if not passive.empty:
            passive["Комментарий"] = "Пассивный счет имеет дебетовое (отрицательное) сальдо"
            self._add("error", "Красное сальдо: пассивный счет с дебетовым остатком", passive)

    # ============ 4.2 Развернутое сальдо ============
    def check_expanded_balance(self) -> None:
        b = self.balances
        # Развернутое сальдо — ошибка на уровне аналитики: у одного контрагента/договора
        # одновременно дебетовый и кредитовый остаток. У активно-пассивных счетов без
        # аналитики одновременные Д/К-остатки нормальны (разные контрагенты).
        both = b[
            (b["Субконто"] != "-")
            & (b["КонецДебет"] > 0)
            & (b["КонецКредит"] > 0)
        ]
        both["Комментарий"] = "По контрагенту/аналитике одновременно дебетовое и кредитовое сальдо"
        self._add("warning", "Развернутое сальдо по аналитике", both)

    # ============ 4.3 Незакрытое сальдо на конец месяца ============
    def check_unclosed_month_end(self) -> None:
        b = self.balances
        closing = b[
            (b["Счет"].map(account_group).isin(self.closing_accounts))
            & ((b["КонецДебет"] > 0) | (b["КонецКредит"] > 0))
        ]
        closing = closing.copy()
        closing["Комментарий"] = "Остаток по закрываемому счету после закрытия месяца"
        self._add(
            "error",
            "Незакрытое сальдо на конец месяца (закрываемые счета)",
            closing
        )

        # Зависшее сальдо: остаток не меняется между периодами
        b2 = b.sort_values(["Счет", "Субконто", "Период"])
        prev_d = b2.groupby(["Счет", "Субконто"])["КонецДебет"].shift(1)
        prev_k = b2.groupby(["Счет", "Субконто"])["КонецКредит"].shift(1)
        stuck = b2[
            ((b2["КонецДебет"] > 0) | (b2["КонецКредит"] > 0))
            & (b2["КонецДебет"] == prev_d)
            & (b2["КонецКредит"] == prev_k)
            & prev_d.notna()
        ]
        stuck = stuck.copy()
        stuck["Комментарий"] = "Сальдо не меняется между периодами (зависший остаток)"
        self._add(
            "warning",
            "Зависшее сальдо (не меняется между периодами)",
            stuck
        )

    # ============ 4.4 Счет 000 ============
    def check_account_000(self) -> None:
        acc = self.balances[
            (self.balances["Счет"].map(account_group) == "000")
            & ((self.balances["КонецДебет"] > 0) | (self.balances["КонецКредит"] > 0))
        ]
        acc = acc.copy()
        acc["Комментарий"] = "Незакрытый остаток на служебном счете 000"
        self._add("error", "Незакрытое сальдо на счете 000", acc)

    # ============ 4.5 Незакрытые расчеты с контрагентами ============
    def _osv_settlement_balance(self) -> pd.Series:
        b = self.balances
        sett = b[b["Счет"].map(account_group).isin(SETTLEMENT_GROUPS)]
        return (
            sett.groupby("Субконто")["КонецДебет"].sum()
            - sett.groupby("Субконто")["КонецКредит"].sum()
        )

    def check_unclosed_settlements(self) -> None:
        if self.documents is None:
            sett = self.balances[
                self.balances["Счет"].map(account_group).isin(SETTLEMENT_GROUPS)]
            dup = sett[
                (sett["Субконто"] != "-")
                & (sett["КонецДебет"] > 0)
                & (sett["КонецКредит"] > 0)
            ]
            dup = dup.copy()
            dup["Комментарий"] = "Возможные незакрытые расчеты: аванс и долг по одному контрагенту"
            self._add(
                "warning",
                "Контрагенты: развернутое сальдо на счетах расчетов (без реестра документов)",
                dup
            )
            return

        rows: list = []
        for name, g in self.documents.groupby("Контрагент"):
            g = g.sort_values("Дата")
            shipments = g.loc[g["ВидНорм"] == "отгрузка", "Сумма"].sum()
            payments = g.loc[g["ВидНорм"].isin(["оплата", "аванс"]), "Сумма"].sum()
            unknown = g.loc[~g["ВидНорм"].isin(["отгрузка", "оплата", "аванс"])]
            open_amount = shipments - payments
            osv_net = float(self._osv_settlement_balance().get(name, 0.0))

            if abs(open_amount) < 1e-6:
                status = "закрыто"
                comment = "Расчеты закрыты документами"
            elif open_amount > 0:
                status = "недоплата"
                comment = f"Отгружено на {shipments:,.2f}, оплачено на {payments:,.2f} — остаток долга"
            else:
                status = "аванс"
                comment = f"Оплачено на {payments:,.2f}, отгрузок на {shipments:,.2f} — аванс не зачтен"

            if unknown is not None and not unknown.empty:
                comment += f"; не распознаны операции: {', '.join(sorted(set(unknown['Вид']))) }"

            if abs(osv_net - open_amount) > 1e-6:
                comment += f"; расхождение с остатком ОСВ ({osv_net:,.2f})"
                status += "+расхождение"

            rows.append({
                "Период": "",
                "Счет": ", ".join(sorted(set(g["Счет"].astype(str)))),
                "Субконто": name,
                "КонецДебет": max(open_amount, 0.0),
                "КонецКредит": max(-open_amount, 0.0),
                "Сумма": abs(open_amount),
                "Комментарий": comment,
            })

        res = pd.DataFrame(
            rows,
            columns=[
                "Период", "Счет", "Субконто",
                "КонецДебет", "КонецКредит", "Сумма",
                "Комментарий"
            ]
        )
        problems = res[res["Сумма"] > 1e-6]
        self._add("error", "Контрагенты: расчеты не закрыты документами", problems.copy())

        mismatches = res[
            res["Комментарий"].str.contains("расхождение", na=False)
        ]
        self._add(
            "warning",
            "Контрагенты: расхождение документов и остатков ОСВ",
            mismatches.copy()
        )

    # ============ Запуск и отчеты ============
    def run_audit(self) -> list[dict]:
        self.errors = []
        self.check_red_balance()
        self.check_expanded_balance()
        self.check_unclosed_month_end()
        self.check_account_000()
        self.check_unclosed_settlements()
        return self.errors

    def summary_df(self) -> pd.DataFrame:
        rows: list[dict[str, Any]] = [
            {
                "Проверка": e["title"],
                "Уровень": e["level"],
                "Строк": len(e["data"]),
                "Сумма": e["amount"]
            } for e in self.errors
        ]
        return pd.DataFrame(
            rows,
            columns=["Проверка", "Уровень", "Строк", "Сумма"]
        )

    def details_df(self) -> pd.DataFrame:
        rows = []
        for e in self.errors:
            for _, r in e["data"].iterrows():
                rows.append({
                    "Проверка": e["title"],
                    "Уровень": e["level"],
                    "Период": r.get("Период", ""),
                    "Счет": r.get("Счет", ""),
                    "Субконто": r.get("Субконто", ""),
                    "Дебет": r.get("КонецДебет", 0.0),
                    "Кредит": r.get("КонецКредит", 0.0),
                    "Сумма": r.get("Сумма", abs(float(r.get("КонецДебет", 0.0) - r.get("КонецКредит", 0.0)))),
                    "Комментарий": r.get("Комментарий", ""),
                })
        return pd.DataFrame(rows, columns=DETAIL_COLUMNS)

    def report(self) -> dict:
        if not self.errors:
            return {
                "status": "ok",
                "status_label": "Успешно",
                "total_flags": 0, "total_amount": 0.0,
                "summary": self.summary_df(),
                "details": self.details_df()
            }
        return {
            "status": "warning",
            "status_label": "Есть ошибки",
            "total_flags": sum(len(e["data"]) for e in self.errors),
            "total_amount": sum(e["amount"] for e in self.errors),
            "summary": self.summary_df(),
            "details": self.details_df()
        }

    def to_excel(self) -> bytes:
        """
        Сводный + детальный отчет с цветовой индикацией (ТЗ п.6, 14)
        """

        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        summary = self.summary_df()
        details = self.details_df()

        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            summary.to_excel(writer, sheet_name="Сводный отчет", index=False)
            details.to_excel(writer, sheet_name="Детальный отчет", index=False)

            wb = writer.book
            header_fill = PatternFill("solid", fgColor="4472C4")
            header_font = Font(color="FFFFFF", bold=True)
            red_fill = PatternFill("solid", fgColor="FFC7CE")
            yellow_fill = PatternFill("solid", fgColor="FFEB9C")

            for sheet, level_col in (("Сводный отчет", 2), ("Детальный отчет", 2)):
                ws = wb[sheet]
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
                for row in ws.iter_rows(min_row=2):
                    level = row[level_col].value
                    if level == "error":
                        fill = red_fill
                    elif level == "warning":
                        fill = yellow_fill
                    else:
                        continue
                    for cell in row:
                        cell.fill = fill

        return buf.getvalue()
