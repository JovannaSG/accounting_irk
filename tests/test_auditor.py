import pandas as pd
import pytest

from core.auditor import AutoAuditor1C, account_group, normalize_balances


def osv(rows, cols=None):
    return pd.DataFrame(rows, columns=cols or [
        "Период", "Счет", "Субконто", "Тип",
        "НачалоДебет", "НачалоКредит", "ОборотДебет", "ОборотКредит",
        "КонецДебет", "КонецКредит",
    ])


# ---------------- 4.1 Красное сальдо ----------------
def test_red_balance_active_negative_only():
    df = osv([
        ["2026-01-31", "50", "Касса", "A", 0, 0, 0, 0, 0, 5000],      # A с К -> ошибка
        ["2026-01-31", "66", "Кредит", "P", 0, 0, 0, 0, 100000, 0],    # P с Д -> ошибка
        ["2026-01-31", "51", "Расчетный", "A", 0, 0, 0, 0, 500000, 0], # норма
        ["2026-01-31", "70", "Зарплата", "P", 0, 0, 0, 0, 0, 120000],   # норма
    ])
    auditor = AutoAuditor1C(df)
    errors = auditor.run_audit()
    red = [e for e in errors if "Красное сальдо" in e["title"]]
    assert len(red) == 2
    accounts = {c for e in red for c in e["data"]["Счет"]}
    assert accounts == {"50", "66"}


def test_red_balance_passive_debit_only():
    # Дебетовый остаток по пассивному счету — красное сальдо (пассив)
    df = osv([
        ["2026-01-31", "68", "Налоги", "P", 0, 0, 0, 0, 5000, 0],
    ])
    auditor = AutoAuditor1C(df)
    errors = auditor.run_audit()
    red = [e for e in errors if "Красное сальдо" in e["title"]]
    assert len(red) == 1
    assert "пассивный счет с дебетовым остатком" in red[0]["title"]
    assert set(red[0]["data"]["Счет"]) == {"68"}


def test_red_balance_net_value_not_raw_columns():
    # Активный счет с КД=1000 и КК=500 -> итог дебетовый, ошибки быть не должно
    df = osv([
        ["2026-01-31", "51", "Расчетный", "A", 0, 0, 0, 0, 1000, 500],
    ])
    auditor = AutoAuditor1C(df)
    errors = auditor.run_audit()
    assert not [e for e in errors if "Красное сальдо" in e["title"]]


def test_ap_account_not_red_by_default():
    df = osv([
        ["2026-01-31", "60.01", "Ромашка", "AP", 0, 0, 0, 0, 45000, 30000],
    ])
    auditor = AutoAuditor1C(df)
    errors = auditor.run_audit()
    assert not [e for e in errors if "Красное сальдо" in e["title"]]


def test_settlement_and_loss_accounts_not_red():
    # 71/73/84 — активно-пассивные в 1С:Бухгалтерия 8.3:
    # перерасход подотчета (К по 71), расчеты по займам (К по 73),
    # непокрытый убыток (Д по 84) — не красное сальдо.
    df = osv([
        ["2026-01-31", "71", "Сидоров", "AP", 0, 0, 0, 0, 0, 1100327],
        ["2026-01-31", "73", "Иванов", "AP", 0, 0, 0, 0, 0, 360],
        ["2026-01-31", "84", "Убыток", "AP", 0, 0, 0, 0, 7118253, 0],
    ])
    auditor = AutoAuditor1C(df)
    errors = auditor.run_audit()
    assert not [e for e in errors if "Красное сальдо" in e["title"]]


def test_red_balance_period_of_occurrence():
    df = osv([
        ["2026-01-31", "50", "Касса", "A", 0, 0, 0, 0, 5000, 0],      # норма
        ["2026-02-28", "50", "Касса", "A", 0, 5000, 0, 0, 0, 5000],   # красное в феврале
    ])
    auditor = AutoAuditor1C(df)
    errors = auditor.run_audit()
    red = [e for e in errors if "Красное сальдо" in e["title"]]
    assert len(red) == 1
    assert "отрицательное сальдо с 28.02.2026" in red[0]["data"].iloc[0]["Комментарий"]


# ---------------- 4.2 Развернутое сальдо ----------------
def test_expanded_balance():
    df = osv([
        ["2026-01-31", "60.01", "Ромашка", "AP", 0, 0, 0, 0, 45000, 30000],
        ["2026-01-31", "62.01", "Смирнов", "AP", 0, 0, 0, 0, 250000, 0],
    ])
    auditor = AutoAuditor1C(df)
    errors = auditor.run_audit()
    exp = [e for e in errors if "Развернутое сальдо" in e["title"]]
    assert len(exp) == 1
    assert list(exp[0]["data"]["Счет"]) == ["60"]


def test_expanded_balance_ignores_account_level_ap():
    # У активно-пассивного счета без аналитики одновременные Д/К-остатки нормальны
    df = osv([
        ["2026-01-31", "60", "-", "AP", 0, 0, 0, 0, 45000, 30000],
    ])
    auditor = AutoAuditor1C(df)
    errors = auditor.run_audit()
    assert not [e for e in errors if "Развернутое сальдо" in e["title"]]


# ---------------- 4.3 Незакрытое сальдо на конец месяца ----------------
def test_unclosed_closing_account():
    df = osv([
        ["2026-01-31", "26", "Общехоз", "A", 0, 0, 0, 0, 120000, 0],
        ["2026-01-31", "44", "Расходы", "A", 0, 0, 0, 0, 0, 0],
        ["2026-01-31", "51", "Расчетный", "A", 0, 0, 0, 0, 100000, 0],
    ])
    auditor = AutoAuditor1C(df)
    errors = auditor.run_audit()
    unclosed = [e for e in errors if "Незакрытое сальдо на конец месяца" in e["title"]]
    assert len(unclosed) == 1
    assert list(unclosed[0]["data"]["Счет"]) == ["26"]


def test_unclosed_period_of_occurrence():
    df = osv([
        ["2026-01-31", "26", "Общехоз", "A", 0, 0, 0, 0, 0, 0],
        ["2026-02-28", "26", "Общехоз", "A", 0, 0, 0, 0, 120000, 0],
    ])
    auditor = AutoAuditor1C(df)
    errors = auditor.run_audit()
    unclosed = [e for e in errors if "Незакрытое сальдо на конец месяца" in e["title"]]
    assert len(unclosed) == 1
    assert "остаток с 28.02.2026" in unclosed[0]["data"].iloc[0]["Комментарий"]


def test_stuck_balance_across_periods():
    df = osv([
        ["2026-01-31", "76.05", "Вектор", "AP", 0, 0, 0, 0, 0, 20000],
        ["2026-02-28", "76.05", "Вектор", "AP", 0, 20000, 0, 0, 0, 20000],
    ])
    auditor = AutoAuditor1C(df, stuck_balance_checks=True)
    errors = auditor.run_audit()
    stuck = [e for e in errors if "Зависшее сальдо" in e["title"]]
    assert len(stuck) == 1
    assert list(stuck[0]["data"]["Счет"]) == ["76"]


def test_stuck_disabled_by_default():
    df = osv([
        ["2026-01-31", "76.05", "Вектор", "AP", 0, 0, 0, 0, 0, 20000],
        ["2026-02-28", "76.05", "Вектор", "AP", 0, 20000, 0, 0, 0, 20000],
    ])
    auditor = AutoAuditor1C(df)
    errors = auditor.run_audit()
    assert not [e for e in errors if "Зависшее сальдо" in e["title"]]


def test_no_stuck_with_single_period():
    df = osv([
        ["2026-01-31", "51", "Расчетный", "A", 0, 0, 0, 0, 500000, 0],
    ])
    auditor = AutoAuditor1C(df)
    errors = auditor.run_audit()
    assert not [e for e in errors if "Зависшее сальдо" in e["title"]]


def test_stuck_series_aggregated_to_one_finding():
    """Три периода с одним остатком — одна находка, а не две строки."""

    df = osv([
        ["2026-01-31", "76.05", "Вектор", "AP", 0, 0, 0, 0, 0, 20000],
        ["2026-02-28", "76.05", "Вектор", "AP", 0, 20000, 0, 0, 0, 20000],
        ["2026-03-31", "76.05", "Вектор", "AP", 0, 20000, 0, 0, 0, 20000],
    ])
    auditor = AutoAuditor1C(df, stuck_balance_checks=True)
    errors = auditor.run_audit()
    stuck = [e for e in errors if "Зависшее сальдо" in e["title"]]
    assert len(stuck) == 1
    data = stuck[0]["data"]
    assert len(data) == 1
    row = data.iloc[0]
    # Период — последний месяц серии, комментарий ссылается на начало
    assert row["Период"] == "2026-03-31"
    assert "31.01.2026" in row["Комментарий"]
    assert "3 пер." in row["Комментарий"]


def test_stuck_streak_break_gives_two_findings():
    """Сальдо зависло, изменилось и зависло снова.
    После схлопывания по родительскому счету на один (счет, субконто)
    остаётся только самая свежая находка (последний период)."""

    df = osv([
        ["2026-01-31", "76.05", "Вектор", "AP", 0, 0, 0, 0, 0, 100],
        ["2026-02-28", "76.05", "Вектор", "AP", 0, 100, 0, 0, 0, 100],
        ["2026-03-31", "76.05", "Вектор", "AP", 0, 0, 50, 0, 0, 150],
        ["2026-04-30", "76.05", "Вектор", "AP", 0, 150, 0, 0, 0, 150],
    ])
    auditor = AutoAuditor1C(df, stuck_balance_checks=True)
    errors = auditor.run_audit()
    stuck = [e for e in errors if "Зависшее сальдо" in e["title"]]
    assert len(stuck) == 1
    data = stuck[0]["data"]
    assert list(data["Период"]) == ["2026-04-30"]
    comments = " | ".join(data["Комментарий"])
    assert "31.03.2026" in comments


def test_stuck_different_subconto_not_merged():
    """Одинаковые остатки у разных субконто — независимые находки."""

    df = osv([
        ["2026-01-31", "62.01", "Ромашка", "AP", 0, 0, 0, 0, 0, 700],
        ["2026-01-31", "62.01", "Лютик", "AP", 0, 0, 0, 0, 0, 700],
        ["2026-02-28", "62.01", "Ромашка", "AP", 0, 0, 0, 0, 0, 700],
        ["2026-03-31", "62.01", "Лютик", "AP", 0, 0, 0, 0, 0, 700],
    ])
    auditor = AutoAuditor1C(df, stuck_balance_checks=True)
    errors = auditor.run_audit()
    stuck = [e for e in errors if "Зависшее сальдо" in e["title"]]
    assert len(stuck) == 1
    assert len(stuck[0]["data"]) == 2


def test_stuck_sorts_periods_chronologically_across_year():
    """Даты «дд.мм.гггг» упорядочиваются по календарю: серия через границу
    года остается одной серией с правильным началом и последним периодом."""

    df = osv([
        ["31.01.2026", "76.05", "Вектор", "AP", 0, 20000, 0, 0, 0, 20000],
        ["28.02.2026", "76.05", "Вектор", "AP", 0, 20000, 0, 0, 0, 20000],
        ["31.12.2025", "76.05", "Вектор", "AP", 0, 0, 0, 0, 0, 20000],
    ])
    auditor = AutoAuditor1C(df, stuck_balance_checks=True)
    errors = auditor.run_audit()
    stuck = [e for e in errors if "Зависшее сальдо" in e["title"]]
    assert len(stuck) == 1
    row = stuck[0]["data"].iloc[0]
    assert row["Период"] == "28.02.2026"
    assert "31.12.2025" in row["Комментарий"]
    assert "3 пер." in row["Комментарий"]


def test_group_balances_disabled_by_default():
    df = osv([
        ["2026-01-31", "51", "Расчетный", "A", 0, 0, 0, 0, 500000, 0],
    ])
    auditor = AutoAuditor1C(df)
    errors = auditor.run_audit()
    assert not [e for e in errors if "Контроль групп счетов" in e["title"]]


def test_group_balances_finds_unclosed_group():
    # Денежные средства (группа 51) не должны иметь остатка на конец периода
    df = osv([
        ["2026-01-31", "51", "Расчетный", "A", 0, 0, 0, 0, 500000, 0],
        ["2026-01-31", "41", "Товары", "A", 0, 0, 0, 0, 100000, 0],
    ])
    auditor = AutoAuditor1C(df, balance_group_checks=True)
    errors = auditor.run_audit()
    groups = [e for e in errors if "Контроль групп счетов" in e["title"]]
    assert len(groups) == 1
    subconto = set(groups[0]["data"]["Субконто"])
    assert "Денежные средства" in subconto
    assert "Товары" in subconto


def test_group_balances_skips_closed_group():
    df = osv([
        ["2026-01-31", "66", "Кредит", "P", 0, 0, 0, 0, 0, 0],
        ["2026-01-31", "51", "Расчетный", "A", 0, 0, 0, 0, 100000, 0],
    ])
    auditor = AutoAuditor1C(df, balance_group_checks=True)
    errors = auditor.run_audit()
    groups = [e for e in errors if "Контроль групп счетов" in e["title"]]
    assert len(groups) == 1
    assert "Кредиты и займы" not in set(groups[0]["data"]["Субконто"])


# ---------------- 4.4 Счет 000 ----------------
def test_account_000_detected_after_string_conversion():
    # Ключевой баг старой версии: '000' превращался в float и проверка не срабатывала
    df = pd.DataFrame([
        {"Период": "2026-01-31", "Счет": "000", "Субконто": "-", "Тип": "AP",
         "НачалоДебет": 0, "НачалоКредит": 0, "ОборотДебет": 1000, "ОборотКредит": 0,
         "КонецДебет": 1000, "КонецКредит": 0},
    ])
    auditor = AutoAuditor1C(df)
    errors = auditor.run_audit()
    acc000 = [e for e in errors if "счете 000" in e["title"]]
    assert len(acc000) == 1
    assert list(acc000[0]["data"]["Счет"]) == ["000"]


# ---------------- 4.5 Контрагенты ----------------
def _docs(rows):
    return pd.DataFrame(rows, columns=["Дата", "Документ", "Контрагент", "Счет", "Вид", "Сумма"])


def test_unclosed_settlements_with_documents():
    balances = osv([
        ["2026-02-28", "62.01", "ИП Смирнов", "AP", 0, 0, 0, 0, 100000, 0],
    ])
    docs = _docs([
        ["2026-01-20", "Отгрузка №3", "ИП Смирнов", "62.01", "отгрузка", 250000],
        ["2026-02-15", "Оплата №3", "ИП Смирнов", "51", "оплата", 150000],
    ])
    auditor = AutoAuditor1C(balances, documents_df=docs)
    errors = auditor.run_audit()
    unclosed = [e for e in errors if "не закрыты документами" in e["title"]]
    assert len(unclosed) == 1
    assert list(unclosed[0]["data"]["Субконто"]) == ["ИП Смирнов"]
    assert unclosed[0]["data"].iloc[0]["Сумма"] == pytest.approx(100000)


def test_closed_settlements_not_flagged():
    balances = osv([
        ["2026-02-28", "62.01", "ИП Смирнов", "AP", 0, 0, 0, 0, 0, 0],
    ])
    docs = _docs([
        ["2026-01-20", "Отгрузка №3", "ИП Смирнов", "62.01", "отгрузка", 250000],
        ["2026-02-15", "Оплата №3", "ИП Смирнов", "51", "оплата", 250000],
    ])
    auditor = AutoAuditor1C(balances, documents_df=docs)
    errors = auditor.run_audit()
    assert not [e for e in errors if "не закрыты документами" in e["title"]]


def test_settlements_heuristic_without_documents():
    balances = osv([
        ["2026-01-31", "60.01", "Ромашка", "AP", 0, 0, 0, 0, 45000, 30000],
    ])
    auditor = AutoAuditor1C(balances)
    errors = auditor.run_audit()
    heuristic = [e for e in errors if "без реестра документов" in e["title"]]
    assert len(heuristic) == 1
    assert list(heuristic[0]["data"]["Субконто"]) == ["Ромашка"]


def test_settlement_advance_consumed_against_debt():
    # Аванс 50 000 зачитывается против отгрузки 100 000 → остаток долга 50 000
    balances = osv([
        ["2026-02-28", "62.01", "ИП Смирнов", "AP", 0, 0, 0, 0, 50000, 0],
    ])
    docs = _docs([
        ["2026-01-20", "Отгрузка №1", "ИП Смирнов", "62.01", "отгрузка", 100000],
        ["2026-01-25", "Аванс №1", "ИП Смирнов", "51", "аванс", 50000],
    ])
    auditor = AutoAuditor1C(balances, documents_df=docs)
    errors = auditor.run_audit()
    unclosed = [e for e in errors if "не закрыты документами" in e["title"]]
    assert len(unclosed) == 1
    row = unclosed[0]["data"].iloc[0]
    assert row["Сумма"] == pytest.approx(50000)
    assert "остаток долга 50 000,00" in row["Комментарий"]
    assert "старейший долг от 20.01.2026" in row["Комментарий"]


def test_settlement_unused_advance_flagged():
    # Только аванс без отгрузок → незачтенный аванс (кредитовый остаток)
    balances = osv([
        ["2026-02-28", "62.02", "ООО Вектор", "AP", 0, 0, 0, 0, 0, 20000],
    ])
    docs = _docs([
        ["2026-01-12", "Аванс №1", "ООО Вектор", "51", "аванс", 20000],
    ])
    auditor = AutoAuditor1C(balances, documents_df=docs)
    errors = auditor.run_audit()
    unclosed = [e for e in errors if "не закрыты документами" in e["title"]]
    assert len(unclosed) == 1
    row = unclosed[0]["data"].iloc[0]
    assert row["Сумма"] == pytest.approx(20000)
    assert row["КонецКредит"] == pytest.approx(20000)
    assert "незачтенный аванс 20 000,00" in row["Комментарий"]


def test_settlement_overpayment_flagged():
    balances = osv([
        ["2026-02-28", "62.01", "ИП Смирнов", "AP", 0, 0, 0, 0, 0, 20000],
    ])
    docs = _docs([
        ["2026-01-20", "Отгрузка №1", "ИП Смирнов", "62.01", "отгрузка", 50000],
        ["2026-02-01", "Оплата №1", "ИП Смирнов", "51", "оплата", 70000],
    ])
    auditor = AutoAuditor1C(balances, documents_df=docs)
    errors = auditor.run_audit()
    unclosed = [e for e in errors if "не закрыты документами" in e["title"]]
    assert len(unclosed) == 1
    row = unclosed[0]["data"].iloc[0]
    assert row["Сумма"] == pytest.approx(20000)
    assert "переплата 20 000,00" in row["Комментарий"]


def test_settlement_aging_oldest_unpaid_shipment():
    balances = osv([
        ["2026-02-28", "62.01", "ИП Смирнов", "AP", 0, 0, 0, 0, 100000, 0],
    ])
    docs = _docs([
        ["2025-11-12", "Отгрузка №1", "ИП Смирнов", "62.01", "отгрузка", 100000],
    ])
    auditor = AutoAuditor1C(balances, documents_df=docs)
    errors = auditor.run_audit()
    unclosed = [e for e in errors if "не закрыты документами" in e["title"]]
    row = unclosed[0]["data"].iloc[0]
    assert "остаток долга 100 000,00" in row["Комментарий"]
    assert "старейший долг от 12.11.2025" in row["Комментарий"]


def test_settlement_osv_account_breakdown_and_split_warning():
    # Долг 60.01 Д и аванс 60.02 К у одного контрагента: по-счетная разбивка + warning
    balances = osv([
        ["2026-02-28", "60.01", "ООО Ромашка", "AP", 0, 0, 0, 0, 45000, 0],
        ["2026-02-28", "60.02", "ООО Ромашка", "AP", 0, 0, 0, 0, 0, 30000],
    ])
    docs = _docs([
        ["2026-01-10", "Аванс №1", "ООО Ромашка", "51", "аванс", 30000],
    ])
    auditor = AutoAuditor1C(balances, documents_df=docs)
    errors = auditor.run_audit()
    split = [e for e in errors if "по разным счетам" in e["title"]]
    assert len(split) == 1
    row = split[0]["data"].iloc[0]
    assert row["Счет"] == "60"
    assert "60.01 Д 45 000,00" in row["Комментарий"]
    assert "60.02 К 30 000,00" in row["Комментарий"]
    unclosed = [e for e in errors if "не закрыты документами" in e["title"]]
    assert unclosed[0]["data"].iloc[0]["Счет"] == "60"


def test_settlement_split_warning_not_for_cross_group():
    # Контрагент одновременно поставщик (60.01 К) и покупатель (62.01 Д) — норма
    balances = osv([
        ["2026-02-28", "60.01", "ООО Вектор", "AP", 0, 0, 0, 0, 0, 50000],
        ["2026-02-28", "62.01", "ООО Вектор", "AP", 0, 0, 0, 0, 30000, 0],
    ])
    auditor = AutoAuditor1C(balances)
    errors = auditor.run_audit()
    assert not [e for e in errors if "по разным счетам" in e["title"]]


def test_settlement_expected_payment_text():
    balances = osv([
        ["2026-02-28", "62.01", "ИП Смирнов", "AP", 0, 0, 0, 0, 100000, 0],
    ])
    docs = _docs([
        ["2025-11-12", "Отгрузка №1", "ИП Смирнов", "62.01", "отгрузка", 100000],
    ])
    auditor = AutoAuditor1C(balances, documents_df=docs)
    errors = auditor.run_audit()
    row = [e for e in errors if "не закрыты документами" in e["title"]][0]["data"].iloc[0]
    assert "остаток долга 100 000,00" in row["Комментарий"]
    assert "старейший долг от 12.11.2025" in row["Комментарий"]


def test_settlement_expected_shipment_text():
    balances = osv([
        ["2026-02-28", "62.02", "ООО Вектор", "AP", 0, 0, 0, 0, 0, 20000],
    ])
    docs = _docs([
        ["2026-01-12", "Аванс №1", "ООО Вектор", "51", "аванс", 20000],
    ])
    auditor = AutoAuditor1C(balances, documents_df=docs)
    errors = auditor.run_audit()
    row = [e for e in errors if "не закрыты документами" in e["title"]][0]["data"].iloc[0]
    assert "незачтенный аванс 20 000,00" in row["Комментарий"]
    assert "старейший незачтенный аванс от 12.01.2026" in row["Комментарий"]


# ---------------- Валидация ----------------
def test_missing_columns_raises():
    with pytest.raises(ValueError, match="Отсутствуют обязательные колонки"):
        AutoAuditor1C(pd.DataFrame({"Счет": ["51"]}))


def test_bad_type_raises():
    df = osv([["2026-01-31", "51", "Расчетный", "X", 0, 0, 0, 0, 100, 0]])
    with pytest.raises(ValueError, match="Недопустимые значения Тип"):
        AutoAuditor1C(df)


def test_non_numeric_raises():
    df = osv([["2026-01-31", "51", "Расчетный", "A", 0, 0, 0, 0, "abc", 0]])
    with pytest.raises(ValueError, match="нечисловые значения"):
        AutoAuditor1C(df)


def test_legacy_two_column_format():
    # Старый формат Счет,Субконто,Тип,Дебет,Кредит
    df = pd.DataFrame([
        {"Счет": "000", "Субконто": "-", "Тип": "AP", "Дебет": 1000.0, "Кредит": 0.0},
    ])
    auditor = AutoAuditor1C(df)
    errors = auditor.run_audit()
    assert [e for e in errors if "счете 000" in e["title"]]


# ---------------- Отчеты ----------------
def test_report_structure_and_excel():
    df = osv([
        ["2026-01-31", "50", "Касса", "A", 0, 0, 0, 0, 0, 5000],
    ])
    auditor = AutoAuditor1C(df)
    auditor.run_audit()
    report = auditor.report()
    assert report["status"] == "warning"
    assert report["total_flags"] == 1
    assert not report["summary"].empty
    assert not report["details"].empty

    xlsx = auditor.to_excel()
    assert xlsx.startswith(b"PK")


def test_report_ok_when_no_errors():
    df = osv([
        ["2026-01-31", "51", "Расчетный", "A", 0, 0, 0, 0, 100000, 0],
    ])
    auditor = AutoAuditor1C(df)
    auditor.run_audit()
    assert auditor.report()["status"] == "ok"


# ---------------- Вспомогательное ----------------
def test_account_group():
    assert account_group("60.01") == "60"
    assert account_group("000") == "000"
    assert account_group("90") == "90"


# ---------------- Переключатели проверок ----------------
def test_checks_subset_disables_others():
    df = osv([
        ["2026-01-31", "50", "Касса", "A", 0, 0, 0, 0, 0, 5000],       # 4.1
        ["2026-01-31", "60.01", "Ромашка", "AP", 0, 0, 0, 0, 45000, 30000],  # 4.2
        ["2026-01-31", "26", "Общехоз", "A", 0, 0, 0, 0, 120000, 0],   # 4.3
        ["2026-01-31", "000", "-", "AP", 0, 0, 0, 0, 1000, 0],         # 4.4
    ])
    auditor = AutoAuditor1C(df, checks={"red_balance", "account_000"})
    errors = auditor.run_audit()
    titles = [e["title"] for e in errors]
    assert any("Красное сальдо" in t for t in titles)
    assert any("счете 000" in t for t in titles)
    assert not any("Развернутое сальдо" in t for t in titles)
    assert not any("закрываемые счета" in t for t in titles)


def test_checks_unknown_key_raises():
    df = osv([["2026-01-31", "51", "Расчетный", "A", 0, 0, 0, 0, 100, 0]])
    with pytest.raises(ValueError, match="Неизвестные ключи: nonexistent"):
        AutoAuditor1C(df, checks={"nonexistent"})


def test_checks_empty_set_runs_nothing():
    df = osv([["2026-01-31", "50", "Касса", "A", 0, 0, 0, 0, 0, 5000]])
    auditor = AutoAuditor1C(df, checks=set())
    assert auditor.run_audit() == []
    assert auditor.report()["status"] == "ok"


# ---------------- Метаданные отчета ----------------
def test_report_meta_keys():
    df = osv([["2026-01-31", "50", "Касса", "A", 0, 0, 0, 0, 0, 5000]])
    auditor = AutoAuditor1C(
        df,
        meta={"organization": "ООО Тест", "period": "Январь 2026", "title": "ОСВ"},
    )
    auditor.run_audit()
    report = auditor.report()
    assert report["organization"] == "ООО Тест"
    assert report["period"] == "Январь 2026"
    assert report["title"] == "ОСВ"


def test_summary_contains_recommendations():
    df = osv([["2026-01-31", "50", "Касса", "A", 0, 0, 0, 0, 0, 5000]])
    auditor = AutoAuditor1C(df)
    auditor.run_audit()
    summary = auditor.summary_df()
    assert "Рекомендации" in summary.columns
    assert summary.iloc[0]["Рекомендации"]  # непустая рекомендация по 4.1


def test_to_pdf():
    pytest.importorskip("fpdf")
    df = osv([["2026-01-31", "50", "Касса", "A", 0, 0, 0, 0, 0, 5000]])
    auditor = AutoAuditor1C(
        df, meta={"organization": "ООО Тест", "period": "Январь 2026"}
    )
    auditor.run_audit()
    data = auditor.to_pdf()
    assert data.startswith(b"%PDF")


def test_to_pdf_multiple_recommendations():
    pytest.importorskip("fpdf")
    rows = [
        ["2026-01-31", "50", "Касса", "A", 0, 0, 0, 0, 0, 5000],
        ["2026-01-31", "000", "Вспомогательный", "A", 0, 0, 0, 0, 1000, 0],
        ["2026-01-31", "26", "Общехоз. расходы", "A", 0, 0, 0, 0, 3000, 0],
    ]
    auditor = AutoAuditor1C(osv(rows))
    auditor.run_audit()
    assert len(auditor.summary_df()) > 1
    data = auditor.to_pdf()
    assert data.startswith(b"%PDF")


def test_excel_overview_sheet_first():
    df = osv([["2026-01-31", "50", "Касса", "A", 0, 0, 0, 0, 0, 5000]])
    auditor = AutoAuditor1C(df, meta={"organization": "ООО Тест"})
    auditor.run_audit()

    import openpyxl
    import io

    wb = openpyxl.load_workbook(io.BytesIO(auditor.to_excel()))
    # «Обзор» — первый лист, старого листа «Об отчете» больше нет
    assert wb.sheetnames[0] == "Обзор"
    assert "Об отчете" not in wb.sheetnames

    ws = wb["Обзор"]
    values = {
        (r[0], r[1])
        for r in ws.iter_rows(min_row=1, max_row=6, values_only=True)
    }
    assert ("Параметр", "Значение") in values
    assert ("Организация", "ООО Тест") in values
    assert ("Статус", "Есть ошибки") in values

    text = "\n".join(
        str(c.value) for row in ws.iter_rows() for c in row if c.value
    )
    assert "Сводка по проверкам" in text
    assert "Крупнейшие нарушения" in text


def test_excel_data_sheets_polished():
    df = osv([
        ["2026-01-31", "50", "Касса", "A", 0, 0, 0, 0, 0, 5000],
        ["2026-02-28", "26", "Общехоз", "A", 0, 0, 0, 0, 120000, 0],
    ])
    auditor = AutoAuditor1C(df)
    auditor.run_audit()

    import openpyxl
    import io

    wb = openpyxl.load_workbook(io.BytesIO(auditor.to_excel()))
    for name in ("Сводный отчет", "Детальный отчет", "По счетам"):
        ws = wb[name]
        assert ws.freeze_panes == "A2", name
        assert ws.auto_filter.ref, name
        assert ws.column_dimensions["A"].width >= 9, name
    # Денежная колонка «Сумма» в детальном отчете — числовой формат
    detail = wb["Детальный отчет"]
    headers = {c.value: c.column for c in detail[1]}
    money_col = headers["Сумма"]
    sample = detail.cell(row=2, column=money_col)
    assert "#" in sample.number_format and "0" in sample.number_format


def test_top_findings_sorted_by_amount():
    df = osv([
        ["2026-01-31", "50", "Касса", "A", 0, 0, 0, 0, 0, 5000],
        ["2026-01-31", "26", "Общехоз", "A", 0, 0, 0, 0, 900000, 0],
        ["2026-01-31", "000", "-", "AP", 0, 0, 1000, 0, 40000, 0],
    ])
    auditor = AutoAuditor1C(df, ml_enabled=False)
    auditor.run_audit()
    top = auditor.top_findings_df(2)
    assert len(top) == 2
    amounts = list(top["Сумма"])
    assert amounts[0] == 900000.0
    assert amounts[1] == 40000.0


def test_normalize_balances_keeps_zero_string_account():
    df = pd.read_csv("data/sample_data.csv", dtype=str)
    norm = normalize_balances(df)
    assert "000" in set(norm["Счет"])
    assert set(norm["Тип"]) <= {"A", "P", "AP"}


# ---------------- Отчет по счету ----------------
def test_accounts_with_errors_and_account_report_df():
    df = osv([
        ["2026-01-31", "50", "Касса", "A", 0, 0, 0, 0, 0, 5000],               # 4.1 -> 50
        ["2026-01-31", "60.01", "Ромашка", "AP", 0, 0, 45000, 0, 45000, 30000],  # 4.2 -> 60.01
        ["2026-01-31", "000", "-", "AP", 0, 0, 1000, 0, 1000, 0],              # 4.4 -> 000
    ])
    auditor = AutoAuditor1C(
        df, checks={"red_balance", "expanded_balance", "account_000"}
    )
    auditor.run_audit()

    accounts = auditor.accounts_with_errors()
    assert "000" in accounts and "50" in accounts and "60" in accounts

    rep = auditor.account_report_df("60")
    assert not rep.empty
    assert set(rep["Счет"]) == {"60"}

    rep_000 = auditor.account_report_df("000")
    assert not rep_000.empty
    assert all("счете 000" in t or "000" in t for t in rep_000["Проверка"])


def test_account_report_df_multiple_accounts_in_cell():
    # В детальном отчете встречаются строки с «Счет» = «51, 62.01»
    df = osv([
        ["2026-01-31", "51", "-", "A", 0, 0, 100000, 0, 100000, 0],
        ["2026-01-31", "62.01", "Долг", "AP", 0, 0, 50000, 0, 50000, 0],
    ])
    docs = pd.DataFrame([
        ["2026-01-31", "Ромашка", "реализация", "60.01", 50000],
        ["2026-01-31", "Ромашка", "оплата", "62.01", 10000],
    ], columns=["Дата", "Контрагент", "Вид", "Счет", "Сумма"])
    auditor = AutoAuditor1C(
        df, documents_df=docs, checks={"settlements"}, ml_enabled=False
    )
    auditor.run_audit()

    # 4.5 строит строки с объединенными счетами контрагента
    cell_values = {str(c) for c in auditor.details_df()["Счет"].dropna()}
    multi = [c for c in cell_values if "," in c]
    assert multi, f"Ожидались составные счета, получены: {cell_values}"

    accounts = auditor.accounts_with_errors()
    assert "60" in accounts and "62" in accounts
    for c in multi:
        for part in c.split(","):
            part = part.strip()
            assert part in accounts
            assert not auditor.account_report_df(part).empty


def test_account_subconto_and_duplicates():
    df = osv([
        ["2026-01-31", "60.01", "ООО Ромашка", "AP", 0, 0, 45000, 0, 0, 45000],
        ["2026-01-31", "60.01", "ООО Ромашка Плюс", "AP", 0, 0, 0, 30000, 30000, 0],
        ["2026-01-31", "51", "-", "A", 0, 0, 100000, 0, 100000, 0],
    ])
    auditor = AutoAuditor1C(df, checks=set(), ml_enabled=False, dup_threshold=70)
    auditor.run_audit()

    assert auditor.account_subconto("60.01") == ["ООО Ромашка", "ООО Ромашка Плюс"]
    assert auditor.account_subconto("51") == []

    dups = auditor.account_subconto_duplicates("60.01")
    assert len(dups) == 1
    row = dups.iloc[0]
    assert {"ООО Ромашка", "ООО Ромашка Плюс"} <= {row["Название А"], row["Название Б"]}
    assert row["Сходство"] >= 70


def test_accounts_summary_df_structure():
    df = osv([
        ["2026-01-31", "50", "Касса", "A", 0, 0, 0, 0, 0, 5000],
        ["2026-01-31", "60.01", "Ромашка", "AP", 0, 0, 45000, 0, 45000, 30000],
        ["2026-01-31", "000", "-", "AP", 0, 0, 1000, 0, 1000, 0],
    ])
    auditor = AutoAuditor1C(
        df, checks={"red_balance", "expanded_balance", "account_000"}, ml_enabled=False
    )
    auditor.run_audit()

    summary = auditor.accounts_summary_df()
    assert not summary.empty
    assert list(summary.columns) == [
        "Счет", "Кол-во нарушений", "Проверки", "Периоды", "Сумма", "Дубли контрагентов"
    ]
    assert set(summary["Счет"]) >= {"000", "50", "60"}
    assert summary["Кол-во нарушений"].sum() == len(auditor.details_df())


def test_to_pdf_brief_blocks():
    """В PDF сначала кратко: сводка по проверкам и топ нарушений."""

    pytest.importorskip("fpdf")
    pytest.importorskip("pypdf")
    import io as _io

    from pypdf import PdfReader

    rows = [
        ["2026-01-31", "50", "Касса", "A", 0, 0, 0, 0, 0, 5000],
        ["2026-01-31", "26", "Общехоз", "A", 0, 0, 0, 0, 900000, 0],
    ]
    auditor = AutoAuditor1C(osv(rows), ml_enabled=False)
    auditor.run_audit()
    reader = PdfReader(_io.BytesIO(auditor.to_pdf()))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    # Новые краткие блоки вместо «Итогов по разделам»
    assert "Сводка по проверкам" in text
    assert "Крупнейшие нарушения" in text
    assert "Итоги по разделам" not in text
    # Самая крупная сумма видна в топе
    assert "900 000,00" in text


def test_excel_contains_by_account_sheet():
    df = osv([["2026-01-31", "50", "Касса", "A", 0, 0, 0, 0, 0, 5000]])
    auditor = AutoAuditor1C(df)
    auditor.run_audit()

    import io
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(auditor.to_excel()))
    assert "По счетам" in wb.sheetnames
    ws = wb["По счетам"]
    headers = [c.value for c in ws[1]]
    assert "Счет" in headers
    assert any(r[0] == "50" for r in ws.iter_rows(min_row=2, values_only=True))


def test_pdf_sections_and_no_by_account():
    pytest.importorskip("fpdf")
    pytest.importorskip("pypdf")
    import io as _io

    from pypdf import PdfReader

    df = osv([["2026-01-31", "50", "Касса", "A", 0, 0, 0, 0, 0, 5000]])
    auditor = AutoAuditor1C(df)
    auditor.run_audit()

    reader = PdfReader(_io.BytesIO(auditor.to_pdf()))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    # Разделы с находками и локализованные уровни
    assert "Красное и развернутое сальдо" in text
    assert "Ошибка" in text
    # Старый дублирующий раздел «Отчет по счетам» удален
    assert "Отчет по счетам" not in text

    clean = osv([["2026-01-31", "51", "Расчетный", "A", 0, 0, 100000, 0, 100000, 0]])
    auditor_ok = AutoAuditor1C(clean)
    auditor_ok.run_audit()
    reader_ok = PdfReader(_io.BytesIO(auditor_ok.to_pdf()))
    text_ok = "\n".join(page.extract_text() or "" for page in reader_ok.pages)
    assert "Нарушений не выявлено" in text_ok


def test_pdf_sections_short_labels_and_no_empty_columns():
    """Объединенные разделы: короткие метки «Проверка», пустые колонки скрыты."""

    advance = AutoAuditor1C.from_findings([
        {
            "title": "Контрагенты: аванс и долг одновременно по разным счетам (ОСВ)",
            "level": "warning",
            "amount": 3000.0,
            "data": [{
                "Счет": "60.02, 62.02",
                "Субконто": 'ООО "Ромашка"', "КонецДебет": 3000.0,
                "КонецКредит": 0.0, "Сумма": 3000.0, "Комментарий": "долг и аванс",
            }],
        },
        {
            "title": "Контрагенты: расхождение документов и остатков ОСВ",
            "level": "error",
            "amount": 1500.0,
            "data": [{
                "Счет": "60.01", "Субконто": 'ООО "Лютик"',
                "КонецДебет": 1500.0, "КонецКредит": 0.0,
                "Сумма": 1500.0, "Комментарий": "документов нет",
            }],
        },
    ])
    sections = {s["title"]: s["df"] for s in advance._sections()}
    settlements = sections["Расчеты с контрагентами"]

    # Короткие метки вместо полных заголовков
    assert set(settlements["Проверка"]) == {"Аванс и долг", "Расхождение с ОСВ"}
    # Ни у одной находки нет непустого Периода -> колонка скрыта
    assert "Период" not in settlements.columns
    assert "Счет" in settlements.columns


def test_pdf_full_text_without_truncation():
    """Длинные комментарии и назначения платежей не обрезаются (ТЗ п.6)."""

    pytest.importorskip("fpdf")
    pytest.importorskip("pypdf")
    import io as _io

    from pypdf import PdfReader

    df = osv([["2026-01-31", "26", "", "A", 0, 0, 80000, 0, 80000, 0]])
    docs = pd.DataFrame([
        [
            "2026-01-10",
            "Платеж №1",
            "ООО Тень",
            "60.01",
            "оплата",
            500000,
            "Обнал по чеку: перевод денежных средств физическим лицом "
            "без подтверждающих документов и договора",
        ],
    ], columns=["Дата", "Документ", "Контрагент", "Счет", "Вид", "Сумма", "Назначение"])
    auditor = AutoAuditor1C(df, docs, ml_enabled=False, nlp_enabled=True)
    auditor.run_audit()

    reader = PdfReader(_io.BytesIO(auditor.to_pdf()))
    text = "\n".join(
        (page.extract_text() or "").replace("\n", "") for page in reader.pages
    )
    # Хвост длинного назначения присутствует целиком — обрезки на 45/80 символов нет
    assert "подтверждающих" in text
    assert "договора" in text


# ---------------- Схлопывание по родительскому счету ----------------
def test_collapse_recurring_error_latest_period_only():
    """Повторяющаяся каждый месяц ошибка −100 схлопывается в одну строку
    с суммой 100 (не 1200), остаётся только самый свежий период."""
    periods = [
        "2026-01-31", "2026-02-28", "2026-03-31", "2026-04-30",
        "2026-05-31", "2026-06-30", "2026-07-31", "2026-08-31",
        "2026-09-30", "2026-10-31", "2026-11-30", "2026-12-31",
    ]
    df = osv([[p, "50", "Касса", "A", 0, 0, 0, 0, 0, 100] for p in periods])
    auditor = AutoAuditor1C(df, checks={"red_balance"})
    errors = auditor.run_audit()
    red = [e for e in errors if "Красное сальдо" in e["title"]]
    assert len(red) == 1
    assert len(red[0]["data"]) == 1
    assert list(red[0]["data"]["Период"]) == ["2026-12-31"]
    assert red[0]["amount"] == pytest.approx(100)


def test_collapse_subaccounts_sum_abs_no_cancel():
    """Долг 60.01 Д 45к и аванс 60.02 К 30к у одного контрагента:
    после схлопывания — счет 60, сумма 75к (модули не сальдируются)."""
    balances = osv([
        ["2026-02-28", "60.01", "ООО Ромашка", "AP", 0, 0, 0, 0, 45000, 0],
        ["2026-02-28", "60.02", "ООО Ромашка", "AP", 0, 0, 0, 0, 0, 30000],
    ])
    docs = _docs([
        ["2026-01-10", "Аванс №1", "ООО Ромашка", "51", "аванс", 30000],
    ])
    auditor = AutoAuditor1C(balances, documents_df=docs)
    errors = auditor.run_audit()
    split = [e for e in errors if "по разным счетам" in e["title"]]
    assert len(split) == 1
    row = split[0]["data"].iloc[0]
    assert row["Счет"] == "60"
    assert row["Сумма"] == pytest.approx(75000)
    assert split[0]["amount"] == pytest.approx(75000)


def test_collapse_dayfirst_keeps_latest_chronologically():
    """Даты в формате «дд.мм.гггг» упорядочиваются по календарю —
    остаётся самый поздний период."""
    df = osv([
        ["31.01.2026", "50", "Касса", "A", 0, 0, 0, 0, 0, 100],
        ["28.02.2026", "50", "Касса", "A", 0, 0, 0, 0, 0, 200],
    ])
    auditor = AutoAuditor1C(df, checks={"red_balance"})
    errors = auditor.run_audit()
    red = [e for e in errors if "Красное сальдо" in e["title"]]
    assert len(red) == 1
    assert list(red[0]["data"]["Период"]) == ["28.02.2026"]
    assert red[0]["amount"] == pytest.approx(200)


def test_collapse_ml_keeps_raw_subaccount():
    """ML-находки — отдельные события, их субсчета не схлопываются."""
    b = osv([
        ["2026-01-31", "51.01", "Расчетный", "A", 0, 0, 1000, 500, 0, 0],
        ["2026-02-28", "51.01", "Расчетный", "A", 0, 0, 10000000, 500, 0, 0],
    ])
    auditor = AutoAuditor1C(b, ml_enabled=True)
    errors = auditor.run_audit()
    jump = [e for e in errors if "ML: резкий скачок оборотов" in e["title"]]
    assert jump
    assert any(str(c).startswith("51.01") for c in jump[0]["data"]["Счет"])
