import io

import pandas as pd
import pytest

from core.account_pass import (
    PASS_DUPLICATES_COLUMNS,
    PASS_SUMMARY_COLUMNS,
    run_account_pass,
)
from core.auditor import AutoAuditor1C


def osv(rows, cols=None):
    return pd.DataFrame(rows, columns=cols or [
        "Период", "Счет", "Субконто", "Тип",
        "НачалоДебет", "НачалоКредит", "ОборотДебет", "ОборотКредит",
        "КонецДебет", "КонецКредит",
    ])


OPTIONS = {
    "checks": ["red_balance", "expanded_balance"],
    "closing_accounts": ["90"],
    "balance_group_checks": False,
    "ml_enabled": True,
    "ml_amount_anomalies": True,
    "ml_turnover_jumps": True,
    "ml_duplicates": True,
    "dup_threshold": 70,
    "anomaly_min_abs": 1000.0,
}


def _fetch_balances(account: str) -> pd.DataFrame:
    if account == "50":
        return osv([
            ["2026-01-31", "50", "-", "A", 0, 0, 0, 0, 0, 5000],  # красное сальдо
        ])
    if account == "60.01":
        return osv([
            ["2026-01-31", "60.01", "ООО Ромашка", "AP", 0, 0, 0, 0, 45000, 30000],
            ["2026-01-31", "60.01", "ООО Ромашка Плюс", "AP", 0, 0, 0, 0, 0, 10000],
        ])
    return pd.DataFrame()  # счёт без данных — ошибка загрузки


def _make_pass(progress=None) -> dict:
    return run_account_pass(
        ["60.01", "50", "99"],
        _fetch_balances,
        OPTIONS,
        progress=progress,
    )


def test_pass_builds_summary_details_duplicates():
    pass_data = _make_pass()

    summary = pass_data["summary_df"]
    assert list(summary.columns) == PASS_SUMMARY_COLUMNS
    assert set(summary["Счет"]) == {"50", "60.01", "99"}

    # 50 — красное сальдо (error), 60.01 — развернутое сальдо (warning), 99 — ошибка загрузки
    by_code = {str(r["Счет"]): r for _, r in summary.iterrows()}
    assert by_code["50"]["Строк нарушений"] >= 1
    assert by_code["50"]["Уровень"] == "error"
    assert by_code["60.01"]["Строк нарушений"] >= 1
    assert by_code["60.01"]["Уровень"] == "warning"
    assert by_code["99"]["Строк нарушений"] == 0
    assert by_code["99"]["Уровень"] == "error"
    assert by_code["99"]["Ошибка"]

    details = pass_data["details_df"]
    assert set(details["Счет"]) == {"50", "60.01"}
    assert any("Красное сальдо" in t for t in details["Проверка"])
    assert any("Развернутое сальдо" in t for t in details["Проверка"])

    dups = pass_data["duplicates_df"]
    assert list(dups.columns) == PASS_DUPLICATES_COLUMNS
    assert len(dups) >= 1
    assert set(dups["Счет"]) == {"60.01"}
    pair = {dups.iloc[0]["Название А"], dups.iloc[0]["Название Б"]}
    assert pair == {"ООО Ромашка", "ООО Ромашка Плюс"}


def test_pass_is_tolerant_to_fetch_errors():
    calls = []

    def flaky(account):
        calls.append(account)
        if account == "51":
            raise ValueError("OData вернул 401")
        return _fetch_balances(account)

    pass_data = run_account_pass(["51", "50"], flaky, OPTIONS)

    summary = {str(r["Счет"]): r for _, r in pass_data["summary_df"].iterrows()}
    assert summary["51"]["Ошибка"]
    assert summary["51"]["Уровень"] == "error"
    assert summary["50"]["Уровень"] == "error"  # несмотря на сбой 51, 50 обработан
    assert set(pass_data["details_df"]["Счет"]) == {"50"}


def test_pass_reports_progress():
    events = []

    def progress(done, total, account):
        events.append((done, total, account))

    _make_pass(progress=progress)

    assert len(events) == 3
    assert events[-1][0] == events[-1][1] == 3
    assert events[0][2] in ("50", "60.01", "99")
    # события идут по одному счёту за раз
    assert sorted(e[1] for e in events) == [3, 3, 3]


def test_pass_empty_accounts():
    pass_data = run_account_pass([], _fetch_balances, OPTIONS)
    assert pass_data["summary_df"].empty
    assert pass_data["details_df"].empty
    assert pass_data["duplicates_df"].empty
    assert pass_data["by_account"] == {}


def test_pass_forwards_stuck_balance_option():
    """Флаг «зависшее сальдо» доходит до аудитора автопрохода по счетам."""

    def fetch(account):
        if account == "76.05":
            return osv([
                ["2026-01-31", "76.05", "Вектор", "AP", 0, 0, 0, 0, 0, 20000],
                ["2026-02-28", "76.05", "Вектор", "AP", 0, 20000, 0, 0, 0, 20000],
            ])
        return pd.DataFrame()

    # Зависшее сальдо вычисляется внутри проверки 4.3, поэтому в проходе
    # должен быть включен и сам чек незакрытого месяца
    options = dict(OPTIONS)
    options["checks"] = OPTIONS["checks"] + ["unclosed_month_end"]
    options["stuck_balance_checks"] = True

    pass_data = run_account_pass(["76.05"], fetch, options)
    details = pass_data["details_df"]
    assert any("Зависшее сальдо" in t for t in details["Проверка"])

    # Без флага той же проверки не возникает
    options["stuck_balance_checks"] = False
    pass_data = run_account_pass(["76.05"], fetch, options)
    details = pass_data["details_df"]
    assert not any("Зависшее сальдо" in t for t in details["Проверка"])


def test_to_excel_contains_pass_sheet():
    pass_data = _make_pass()
    df = osv([["2026-01-31", "50", "-", "A", 0, 0, 0, 0, 0, 5000]])
    auditor = AutoAuditor1C(df)
    auditor.run_audit()

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(auditor.to_excel(account_pass=pass_data)))
    assert "Проход по счетам" in wb.sheetnames
    ws = wb["Проход по счетам"]
    headers = [c.value for c in ws[1]]
    assert "Счет" in headers
    col = headers.index("Счет")
    assert any(r[col] in ("50", "60.01") for r in ws.iter_rows(min_row=2, values_only=True))


def test_to_pdf_contains_pass_section():
    pytest.importorskip("fpdf")
    pytest.importorskip("pypdf")
    import io as _io

    from pypdf import PdfReader

    pass_data = _make_pass()
    df = osv([["2026-01-31", "50", "-", "A", 0, 0, 0, 0, 0, 5000]])
    auditor = AutoAuditor1C(df)
    auditor.run_audit()

    reader = PdfReader(_io.BytesIO(auditor.to_pdf(account_pass=pass_data)))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    assert "Автопроход по счетам" in text
    # Дублирующий раздел «Отчет по счетам» удален из печатного отчета
    assert "Отчет по счетам" not in text
    assert "Стр." in text  # нумерация страниц
